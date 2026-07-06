from __future__ import annotations

import csv
import io
import zipfile
from dataclasses import dataclass
from typing import Iterable, List, Optional, Tuple

from app.models.part import Part
from app.services.part_norm import clean_rev
from app.services.timezone_utils import format_display_ts, utc_now


@dataclass
class CompileRow:
    part_number: str
    revision: str
    qty: float
    status: str  # "ok" | "missing"


def _clean_rev(value: object) -> str:
    return clean_rev(value)


def _to_qty(value: object) -> float:
    if value is None or str(value).strip() == "":
        return 1.0
    try:
        return float(value)
    except Exception:
        return 1.0


def parse_compile_excel(path: str) -> List[CompileRow]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = None
    for name in wb.sheetnames:
        if name.strip().lower() == "compile":
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c or "").strip().lower() for c in rows[0]]
    def find_col(names: Iterable[str]) -> Optional[int]:
        for name in names:
            if name in header:
                return header.index(name)
        return None

    idx_pn = find_col(["partnumber", "part number", "part_number", "pn"])
    idx_rev = find_col(["revision", "rev"])
    idx_qty = find_col(["qty", "quantity"])

    if idx_pn is None:
        return []

    out: List[CompileRow] = []
    for row in rows[1:]:
        if row is None:
            continue
        pn = str(row[idx_pn] or "").strip()
        if not pn:
            continue
        rev = _clean_rev(row[idx_rev]) if idx_rev is not None else ""
        qty = _to_qty(row[idx_qty]) if idx_qty is not None else 1.0
        out.append(CompileRow(part_number=pn, revision=rev, qty=qty, status="ok"))
    return out


def _expand_with_children(
    resolved: List[Tuple[Part, float]],
    *,
    expand_subassemblies: bool = True,
) -> List[Tuple[str, str, float]]:
    """Turn the uploaded rows into the same fully-expanded flat BOM a real Doc
    Pack would walk to. Each row stands in for a top-level BOM line: if it is
    itself an assembly with its own children in the system, those children
    (and their own files) get pulled in too, scaled by the row's quantity -
    exactly like exporting a real assembly at full depth.
    """
    from app.services.docpacks import _flatten_bom, _norm_rev

    agg: dict = {}

    def add(pn: str, rev: str, qty: float) -> None:
        key = (pn, _norm_rev(rev))
        agg[key] = agg.get(key, 0.0) + float(qty or 0.0)

    for part, qty in resolved:
        pn, rev = part.part_number, part.revision or ""
        add(pn, rev, qty)
        if expand_subassemblies:
            for cpn, crev, cqty in _flatten_bom(pn, rev, full=True):
                add(cpn, crev, cqty * qty)

    return [(pn, rev, q) for (pn, rev), q in agg.items()]


def build_excel_compile_zip(
    rows: List[CompileRow],
    *,
    input_filename: str,
    input_bytes: bytes,
    file_root: str,
    file_groups: Optional[List[str]] = None,
    processes: Optional[List[str]] = None,
    process_mode: str = "all",
    classified_filter: str = "show",
    expand_subassemblies: bool = True,
    want_excel_bom: bool = True,
    excel_all_fields: bool = False,
    excel_field_ids: Optional[List[str]] = None,
    want_pdf_binder: bool = False,
    want_index_pdf: bool = False,
    want_visual_list: bool = False,
    want_hardware_summary: bool = False,
    want_cover_page: bool = False,
    binder_add_cover: bool = True,
    binder_add_index: bool = True,
    binder_add_visual_list: bool = True,
    binder_add_hardware_summary: bool = True,
    binder_add_datasheets: bool = False,
    binder_page_numbers: bool = True,
    binder_include_flat_patterns: bool = False,
    stamp_quote: bool = False,
    stamp_confidential: bool = False,
    stamp_approved: bool = False,
    stamp_wip: bool = False,
    stamp_inprogress: bool = False,
    title: Optional[str] = None,
    description: Optional[str] = None,
    thumbnail_path: Optional[str] = None,
) -> Tuple[bytes, List[CompileRow]]:
    from app.services.docpacks import _part_by

    missing: List[CompileRow] = []
    resolved: List[Tuple[Part, float]] = []

    for r in rows:
        # Same tolerant match used by the doc pack pipeline itself (case,
        # whitespace, ".0"-suffix): an uploaded sheet's PartNumber/Revision
        # text routinely differs slightly in formatting from the canonical
        # DB record, and a strict exact match would wrongly flag it missing.
        part = _part_by(r.part_number, r.revision)
        if not part:
            r.status = "missing"
            missing.append(r)
            continue
        resolved.append((part, r.qty))

    parts = _expand_with_children(resolved, expand_subassemblies=expand_subassemblies)

    buf = io.BytesIO()
    z = zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED)
    z.writestr("input.xlsx", input_bytes)

    # Summary CSV
    csv_buf = io.StringIO()
    w = csv.writer(csv_buf)
    w.writerow(["part_number", "revision", "qty", "status"])
    for r in rows:
        w.writerow([r.part_number, r.revision, f"{r.qty:g}", r.status])
    z.writestr("parts.csv", csv_buf.getvalue())

    if missing:
        msg = "\n".join(f"{r.part_number}\t{r.revision}" for r in missing)
        z.writestr("missing_parts.txt", msg)

    # Delegate everything else (BOM sheet, selected files, PDF binder, index,
    # visual list, hardware summary, cover page, stamps...) to the same
    # Doc Pack pipeline used for real assemblies. The uploaded compile sheet
    # stands in for a BOM tree via flat_override, so it goes through the exact
    # same, already-tested build as any other doc pack.
    if parts:
        from app.services.docpacks import DocPackOptions, build_docpack
        build_ts = utc_now()
        title = (title or "").strip() or f"BOM {format_display_ts(build_ts, fmt='%Y-%m-%d')}"
        opts = DocPackOptions(
            root_pn=title,
            root_rev=None,
            root_desc_override=(description or "").strip(),
            root_thumb_override=thumbnail_path,
            build_ts=build_ts,
            flat_override=parts,
            processes=processes,
            process_mode=process_mode,
            classified_filter=classified_filter,
            file_types=file_groups or ["pdf", "dxf", "step", "datasheet"],
            want_selected_files=True,
            want_excel_bom=want_excel_bom,
            excel_all_fields=excel_all_fields,
            excel_field_ids=excel_field_ids,
            want_pdf_binder=want_pdf_binder,
            want_index_pdf=want_index_pdf,
            want_visual_list=want_visual_list,
            want_hardware_summary=want_hardware_summary,
            want_cover_page=want_cover_page,
            binder_add_cover=binder_add_cover,
            binder_add_index=binder_add_index,
            binder_add_visual_list=binder_add_visual_list,
            binder_add_hardware_summary=binder_add_hardware_summary,
            binder_add_datasheets=binder_add_datasheets,
            binder_page_numbers=binder_page_numbers,
            binder_include_flat_patterns=binder_include_flat_patterns,
            stamp_quote=stamp_quote,
            stamp_confidential=stamp_confidential,
            stamp_approved=stamp_approved,
            stamp_wip=stamp_wip,
            stamp_inprogress=stamp_inprogress,
            output_name=title,
        )
        try:
            _, data, mime = build_docpack(opts)
            if mime == "application/zip":
                inner = zipfile.ZipFile(io.BytesIO(data))
                for name in inner.namelist():
                    z.writestr(name, inner.read(name))
            else:
                ext = "pdf" if "pdf" in mime else "xlsx"
                z.writestr(f"DocPack.{ext}", data)
        except RuntimeError as exc:
            z.writestr("docpack_error.txt", str(exc))

    z.close()
    return buf.getvalue(), missing

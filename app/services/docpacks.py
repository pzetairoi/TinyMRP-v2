from __future__ import annotations
import io, os, re, tempfile, zipfile
from datetime import datetime
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple, Set

from flask import current_app, request, url_for
from urllib.parse import quote
from app.models.part import Part
from app.models.bom import BOMLink
from app.models.artifact import PartFile
from app.services.attrs import harvest_part_attrs, ALIASES, approved_value
from app.services.canonical_fields import canonical_processes_for_part
from app.services.field_config import context_field_ids, get_field_config, resolve_part_field_values
from app.services.processmeta import normalize_processes
from app.services.filenames import build_output_name
from app.services.part_norm import clean_rev as _clean_rev, clean_rev_or_none as _rev_or_none
from app.services.app_settings import format_display_ts, get_display_dt, resolve_brand_logo_path
from app.services.insights import classify_part
from app.services.markup_documents import combine_markup_documents, markup_documents_for_pairs


@dataclass
class DocPackOptions:
    root_pn: str
    root_rev: Optional[str] = None
    depth: str = "full"                 # "top" | "full"
    include_consumed: bool = False       # placeholder (future)
    classified_filter: str = "show"      # "hide" | "show" | "only"
    processes: List[str] = None          # selected process names (canonical)
    process_mode: str = "all"            # "selected" | "all"
    file_types: List[str] = None         # ext_groups to include (e.g., ["png","pdf","step","dxf","edr","3mf","datasheet"]) 
    want_excel_bom: bool = False
    excel_all_fields: bool = False
    excel_field_ids: Optional[List[str]] = None
    want_selected_files: bool = True
    want_pdf_binder: bool = False
    want_index_pdf: bool = False
    want_visual_list: bool = False
    want_cover_page: bool = False
    want_whereused_report: bool = False
    want_hardware_summary: bool = False
    want_markup_files: bool = False
    want_markup_report: bool = False
    fabrication_pack: bool = False
    binder_add_cover: bool = True
    binder_add_index: bool = True
    binder_add_datasheets: bool = False
    binder_add_visual_list: bool = True
    binder_add_hardware_summary: bool = True
    binder_add_whereused: bool = False
    binder_page_numbers: bool = True
    binder_include_flat_patterns: bool = False
    binder_add_markups: bool = False
    # binder stamps
    stamp_quote: bool = False
    stamp_confidential: bool = False
    stamp_approved: bool = False
    stamp_wip: bool = False
    stamp_inprogress: bool = False
    output_name: Optional[str] = None
    build_ts: Optional[datetime] = None
    # When set, used in place of a real BOM tree walk (root_pn/root_rev become a
    # display label only). Lets an ad-hoc parts list (e.g. an uploaded compile
    # sheet) go through the exact same doc pack pipeline as a real assembly.
    flat_override: Optional[List[Tuple[str, str, float]]] = None
    # Fallback description shown on cover/index/binder when root_pn isn't a real
    # part (i.e. there's no Part record to pull a description from).
    root_desc_override: Optional[str] = None
    # Absolute path to an image file used as the cover/visual-list picture for
    # the root when there's no real Part record to pull a thumbnail from
    # (e.g. an uploaded compile sheet's synthetic root).
    root_thumb_override: Optional[str] = None


def _norm_rev(rev: Optional[str]) -> str:
    return _clean_rev(rev)


def _part_by(pn: str, rev: Optional[str]) -> Optional[Part]:
    if rev is None:
        return Part.objects(part_number__iexact=pn).order_by("-updated_at").first()
    # Compare on cleaned values in Python rather than a DB-level exact/iexact
    # match: real-world revision strings routinely carry stray whitespace or
    # formatting differences (e.g. "1 " vs "1") between BOMLink.child_rev and
    # the part's own canonical revision, which a DB-level match would miss.
    target = _clean_rev(rev).lower()
    for p in Part.objects(part_number__iexact=pn):
        if _clean_rev(getattr(p, "revision", "")).lower() == target:
            return p
    return None


def _part_description(part: Optional[Part], attrs: Optional[Dict] = None) -> str:
    if attrs is None and part:
        try:
            attrs = harvest_part_attrs(part)
        except Exception:
            attrs = {}
    try:
        values = resolve_part_field_values(part, ["description"], attrs=attrs, config=get_field_config())
        return str(values.get("description") or "")
    except Exception:
        pass
    if part and getattr(part, "description", ""):
        return str(part.description or "")
    return str((attrs or {}).get("description") or "")


def _part_processes(part: Optional[Part]) -> List[str]:
    if not part:
        return []
    meta = current_app.config.get("PROCESS_META", {}) or {}
    return canonical_processes_for_part(part, process_meta=meta)


def _flatten_bom(
    root_pn: str,
    root_rev: Optional[str],
    full: bool,
    *,
    include_consumed: bool = True,
    terminal_processes: Optional[Iterable[str]] = None,
) -> List[Tuple[str, str, float]]:
    """Return a flat list of (pn, rev, qty). If full=False, only top-level children.
    Accumulates quantities when a child appears multiple times along different paths.
    """
    if root_rev is None:
        root_part = _part_by(root_pn, None)
        root_rev = _clean_rev(getattr(root_part, "revision", "") if root_part else "")
    else:
        root_rev = _clean_rev(root_rev)
    out: Dict[Tuple[str,str], float] = {}

    def add(pn: str, rev: str, qty: float):
        key = (pn, _clean_rev(rev))
        out[key] = out.get(key, 0.0) + float(qty or 0.0)

    terminals: Set[str] = set([str(x).strip().lower() for x in (terminal_processes or [])])

    def children(pn: str, rev: Optional[str]):
        rev_val = _rev_or_none(rev)
        if "parent_rev" in BOMLink._fields:
            if rev_val is not None:
                return BOMLink.objects(parent_pn=pn, parent_rev=rev_val)
        return BOMLink.objects(parent_pn=pn)

    stack: List[Tuple[str,str,float]] = []
    # Seed stack with immediate children
    for l in children(root_pn, root_rev):
        stack.append((l.child_pn, _clean_rev(getattr(l, "child_rev", "") or ""), float(getattr(l, "qty", 1.0) or 1.0)))
        if not full:
            add(l.child_pn, _clean_rev(getattr(l, "child_rev", "") or ""), float(getattr(l, "qty", 1.0) or 1.0))

    if not full:
        return [(pn, rev, qty) for (pn, rev), qty in out.items()]

    while stack:
        pn, rev, q = stack.pop()
        add(pn, rev, q)
        # If we hide consumed components, stop traversing at terminal process parts
        if not include_consumed:
            pdoc = _part_by(pn, rev)
            procs = set(_part_processes(pdoc))
            if terminals and (procs & terminals):
                # do not traverse into its children
                continue
        for l in children(pn, rev):
            cq = float(getattr(l, "qty", 1.0) or 1.0) * q
            stack.append((l.child_pn, _clean_rev(getattr(l, "child_rev", "") or ""), cq))

    return [(pn, rev, qty) for (pn, rev), qty in out.items()]


def _passes_classified_filter(part: Part, mode: str) -> bool:
    mode = (mode or "show").lower()
    attrs = getattr(part, "attrs", {}) or {}
    val = str(attrs.get("classified", "")).strip().lower()
    is_classified = (val in ("yes","true","1","classified","confidential")) or (val not in ("no", "", "false", "0"))
    if mode == "hide":
        return not is_classified
    if mode == "only":
        return is_classified
    return True


def _passes_process_filter(part: Part, selected: Optional[List[str]], mode: str) -> bool:
    if (mode or "all").lower() != "selected":
        return True
    if not selected:
        return True
    procs = _part_processes(part)
    for p in selected:
        if p.lower() in procs:
            return True
    return False


def _collect_files(pn_rev_qty: Iterable[Tuple[str,str,float]], file_types: Optional[List[str]]) -> List[PartFile]:
    groups = set([g.lower() for g in (file_types or [])])
    out: List[PartFile] = []
    for pn, rev, _ in pn_rev_qty:
        q = PartFile.objects(part_number__iexact=pn)
        if rev is not None:
            q = q.filter(revision__iexact=_clean_rev(rev))
        if groups:
            q = q.filter(ext_group__in=list(groups))
        out.extend(list(q))
    return out


def _file_coverage_report(flat: Iterable[Tuple[str, str, float]], file_root: str) -> str:
    """Per-part diagnostic written into every doc pack zip: does a PDF record
    exist for this part, and does the file it points to actually exist on
    disk? Makes gaps in the binder/selected-files bundle visible immediately
    instead of requiring guesswork against a live database.
    """
    file_root = (file_root or "").rstrip("/\\")

    def _status(pf: Optional[PartFile]) -> str:
        if not pf:
            return "no record"
        rel = pf.rel_path.replace("\\", "/") if pf.rel_path else ""
        abs_path = pf.path if os.path.isabs(pf.path) else os.path.join(file_root, rel.replace("/", os.sep))
        return "OK" if os.path.isfile(abs_path) else f"record exists but file missing on disk ({abs_path})"

    lines = ["part_number\trevision\tpdf\tany_file"]
    for pn, rev, _qty in sorted(flat, key=lambda t: (t[0] or "", t[1] or "")):
        qs = PartFile.objects(part_number__iexact=pn, revision__iexact=_clean_rev(rev))
        lines.append(f"{pn}\t{rev}\t{_status(qs.filter(ext__iexact='pdf').first())}\t{_status(qs.first())}")
    return "\n".join(lines)


def _missing_pdf_rows(pairs: Iterable[Tuple[str, str]], file_root: str) -> List[Dict[str, object]]:
    """Parts whose PDF *record* exists but whose file is absent on disk.

    Unlike a missing record (normal for hardware/no-drawing parts), this means
    a page that should be in the binder body silently isn't -- e.g. an upload
    that failed to land on the deliverables volume (permission/ownership
    mismatch) while the DB row was still created. Surfaced so the gap is
    visible in the binder itself instead of only in the zip-only coverage txt.
    """
    file_root = (file_root or "").rstrip("/\\")
    rows: List[Dict[str, object]] = []
    seen: Set[Tuple[str, str]] = set()
    for pn, rev in pairs:
        key = (pn, _norm_rev(rev))
        if key in seen:
            continue
        seen.add(key)
        for pf in PartFile.objects(part_number__iexact=pn, revision__iexact=_clean_rev(rev), ext__iexact="pdf"):
            rel = pf.rel_path.replace("\\", "/") if pf.rel_path else ""
            abs_path = pf.path if os.path.isabs(pf.path) else os.path.join(file_root, rel.replace("/", os.sep))
            if not os.path.isfile(abs_path):
                pdoc = _part_by(pn, rev)
                rows.append({
                    "partnumber": pn,
                    "revision": _norm_rev(rev),
                    "description": _part_description(pdoc) if pdoc else "",
                })
                break
    return rows


def _missing_files_pdf(rows: List[Dict[str, object]]) -> Optional[bytes]:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet
    except Exception:
        return None
    if not rows:
        return None

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    flowables = [
        Paragraph("Missing Source Documents", styles["Heading2"]),
        Spacer(0, 3*mm),
        Paragraph(
            "The parts below have a PDF on record but the file could not be found on the "
            "deliverables volume. Their drawings are NOT included in this binder. Re-export/"
            "re-upload these parts and rebuild the docpack.",
            styles["Normal"],
        ),
        Spacer(0, 4*mm),
    ]
    table_data = [["Part Number", "Revision", "Description"]]
    for r in rows:
        table_data.append([str(r.get("partnumber") or ""), str(r.get("revision") or ""), str(r.get("description") or "")])
    table = Table(table_data, colWidths=[40*mm, 20*mm, 100*mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#b30000")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
    ]))
    flowables.append(table)
    doc.build(flowables)
    return buf.getvalue()


_FLAT_PATTERN_RE = re.compile(r"(?:^|[^a-z0-9])(flatpattern|flat[-_ ]pattern|fp)(?:[^a-z0-9]|$)", re.IGNORECASE)
_FP_LABEL_STRIP_RE = re.compile(r"[^a-z0-9]+", re.IGNORECASE)


def _is_flat_pattern_name(name: str) -> bool:
    base = os.path.splitext(os.path.basename(name or ""))[0].lower()
    if not base:
        return False
    if "flatpattern" in base or "flat-pattern" in base or "flat_pattern" in base or "flat pattern" in base:
        return True
    return bool(_FLAT_PATTERN_RE.search(base))


def _is_flat_pattern_file(pf: PartFile) -> bool:
    name = pf.rel_path or pf.path or ""
    return _is_flat_pattern_name(name)


def _normalize_fp_label(text: str) -> str:
    return _FP_LABEL_STRIP_RE.sub("", (text or "").lower())


def _flat_pattern_page_tokens() -> List[str]:
    cfg = current_app.config.get("FLAT_PATTERN_PAGE_NAMES", None)
    if cfg is None:
        cfg = ["flatpattern"]
    raw: List[str] = []
    if isinstance(cfg, str):
        raw = [p for p in re.split(r"[,;\r\n]+", cfg) if p.strip()]
    else:
        for item in cfg:
            if item is None:
                continue
            raw.append(str(item))
    tokens: List[str] = []
    seen = set()
    for item in raw:
        token = _normalize_fp_label(item)
        if not token or token in seen:
            continue
        tokens.append(token)
        seen.add(token)
    return tokens


def _is_flat_pattern_page_label(label: str, tokens: List[str]) -> bool:
    norm = _normalize_fp_label(label)
    if not norm:
        return False
    for t in tokens:
        if t and t in norm:
            return True
    return False


def _int_to_roman(num: int) -> str:
    if num <= 0:
        return ""
    vals = [
        (1000, "M"),
        (900, "CM"),
        (500, "D"),
        (400, "CD"),
        (100, "C"),
        (90, "XC"),
        (50, "L"),
        (40, "XL"),
        (10, "X"),
        (9, "IX"),
        (5, "V"),
        (4, "IV"),
        (1, "I"),
    ]
    out = []
    n = num
    for v, s in vals:
        while n >= v:
            out.append(s)
            n -= v
    return "".join(out)


def _int_to_letters(num: int) -> str:
    if num <= 0:
        return ""
    out = []
    n = num
    while n > 0:
        n -= 1
        out.append(chr(ord("A") + (n % 26)))
        n //= 26
    return "".join(reversed(out))


def _format_page_label(style: object, num: int) -> str:
    if not style:
        return ""
    style_key = str(style)
    if style_key == "/D":
        return str(num)
    if style_key == "/R":
        return _int_to_roman(num)
    if style_key == "/r":
        return _int_to_roman(num).lower()
    if style_key == "/A":
        return _int_to_letters(num)
    if style_key == "/a":
        return _int_to_letters(num).lower()
    return str(num)


def _collect_page_labels_tree(node: object) -> Dict[int, object]:
    out: Dict[int, object] = {}
    if node is None:
        return out
    try:
        node = node.get_object()
    except Exception:
        pass
    try:
        nums = node.get("/Nums")
        if nums:
            for i in range(0, len(nums), 2):
                try:
                    idx = int(nums[i])
                except Exception:
                    continue
                out[idx] = nums[i + 1]
    except Exception:
        pass
    try:
        kids = node.get("/Kids")
        if kids:
            for kid in kids:
                out.update(_collect_page_labels_tree(kid))
    except Exception:
        pass
    return out


def _page_labels_from_pagelabels(reader) -> List[str]:
    try:
        root = reader.trailer["/Root"]
    except Exception:
        return []
    try:
        labels_root = root.get("/PageLabels")
    except Exception:
        labels_root = None
    if not labels_root:
        return []
    mapping = _collect_page_labels_tree(labels_root)
    if not mapping:
        return []
    total = len(reader.pages)
    labels = [""] * total
    entries = sorted(mapping.items(), key=lambda item: item[0])
    for i, (start_idx, info) in enumerate(entries):
        if start_idx >= total:
            continue
        try:
            info = info.get_object()
        except Exception:
            pass
        try:
            prefix = str(info.get("/P", "")) if info else ""
        except Exception:
            prefix = ""
        try:
            style = info.get("/S") if info else None
        except Exception:
            style = None
        try:
            start_num = int(info.get("/St", 1)) if info else 1
        except Exception:
            start_num = 1
        end_idx = entries[i + 1][0] if i + 1 < len(entries) else total
        end_idx = min(end_idx, total)
        for page_idx in range(start_idx, end_idx):
            label_num = _format_page_label(style, start_num + (page_idx - start_idx))
            label = f"{prefix}{label_num}" if prefix or label_num else ""
            if label:
                labels[page_idx] = label
    return labels


def _page_labels_from_outline(reader) -> List[str]:
    try:
        outlines = reader.outline
    except Exception:
        try:
            outlines = reader.outlines
        except Exception:
            return []
    total = len(reader.pages)
    labels = [""] * total

    def _walk(items):
        for item in items or []:
            if isinstance(item, list):
                _walk(item)
                continue
            title = None
            try:
                title = getattr(item, "title", None)
            except Exception:
                title = None
            if title is None:
                try:
                    title = item.get("/Title")
                except Exception:
                    title = None
            page_num = None
            try:
                page_num = reader.get_destination_page_number(item)
            except Exception:
                try:
                    page_num = reader.getDestinationPageNumber(item)
                except Exception:
                    page_num = None
            if title is None or page_num is None:
                continue
            try:
                idx = int(page_num)
            except Exception:
                continue
            if 0 <= idx < total and not labels[idx]:
                labels[idx] = str(title)

    _walk(outlines)
    if any(labels):
        return labels
    return []


def _page_labels(reader) -> List[str]:
    total = len(reader.pages)
    labels = _page_labels_from_pagelabels(reader)
    if labels and len(labels) < total:
        labels.extend([""] * (total - len(labels)))
    if not labels:
        labels = [""] * total
    outlines = _page_labels_from_outline(reader)
    if outlines:
        for i, label in enumerate(outlines):
            if label and (i >= len(labels) or not labels[i]):
                if i >= len(labels):
                    labels.append(label)
                else:
                    labels[i] = label
    return labels if any(labels) else []


def _pdf_items_for_pairs(
    pairs: Iterable[Tuple[str, str]],
    *,
    include_datasheets: bool,
    include_flat_patterns: bool,
) -> Tuple[List[Dict[str, object]], List[str]]:
    root_dir = (current_app.config.get("FILE_ROOT_LOCAL") or "").rstrip("/\\")
    items: List[Dict[str, object]] = []
    paths: List[str] = []
    for pn, rev in pairs:
        q = PartFile.objects(part_number__iexact=pn)
        if rev is not None:
            q = q.filter(revision__iexact=(rev or ""))
        files = [f for f in q.filter(ext__iexact="pdf")]
        files.sort(key=lambda f: os.path.basename(f.rel_path or f.path or "").lower())
        for f in files:
            if (getattr(f, "ext_group", "") or "").lower() == "datasheet" and not include_datasheets:
                continue
            if not include_flat_patterns and _is_flat_pattern_file(f):
                continue
            rp = f.rel_path.replace("\\", "/") if f.rel_path else os.path.basename(f.path)
            abs_path = f.path if os.path.isabs(f.path) else os.path.join(root_dir, rp.replace("/", os.sep))
            if os.path.isfile(abs_path):
                paths.append(abs_path)
                items.append({
                    "pn": pn,
                    "rev": _norm_rev(rev),
                    "path": abs_path,
                    "ext_group": (getattr(f, "ext_group", "") or "").lower(),
                })
    return items, paths


def _bom_occurrences(
    root_pn: str,
    root_rev: Optional[str],
    *,
    include_consumed: bool = True,
    terminal_processes: Optional[Iterable[str]] = None,
) -> Dict[Tuple[str,str], List[Tuple[str, float]]]:
    """Return mapping (pn,rev) -> list of (level_path, qty_at_that_level).
    Level path follows legacy-style "+.01.02" when sequence metadata exists.
    Applies the same consumed/terminal filtering as the visual list/BOM flatten.
    """
    if root_rev is None:
        root_part = _part_by(root_pn, None)
        root_rev = _clean_rev(getattr(root_part, "revision", "") if root_part else "")
    else:
        root_rev = _clean_rev(root_rev)
    occ: Dict[Tuple[str,str], List[Tuple[str,float]]] = {}
    terminals: Set[str] = set([str(x).strip().lower() for x in (terminal_processes or [])])

    def children(pn: str, rev: Optional[str]):
        rev_val = _rev_or_none(rev)
        if "parent_rev" in BOMLink._fields:
            if rev_val is not None:
                return BOMLink.objects(parent_pn=pn, parent_rev=rev_val)
        return BOMLink.objects(parent_pn=pn)

    def _fmt_seq(seq: object, fallback: int | None) -> str:
        val = seq if seq not in (None, "") else fallback
        if val is None:
            return ""
        try:
            n = int(str(val).strip())
            return f"{n:02d}" if n < 10 else str(n)
        except Exception:
            return str(val).strip()

    def _append_path(base: str, seq: object, fallback: int | None) -> str:
        seg = _fmt_seq(seq, fallback)
        if not seg:
            return base
        return f"{base}.{seg}" if base else seg

    def _link_occurrences(link: BOMLink) -> List[Tuple[float, object]]:
        occs = getattr(link, "occurrences", None) or []
        if occs:
            out = []
            for occ in occs:
                qty_val = occ.get("qty")
                if qty_val is None:
                    qty_val = getattr(link, "qty", 1.0)
                qty = float(qty_val or 0.0)
                out.append((qty, occ.get("seq")))
            return out
        link_qty = getattr(link, "qty", None)
        if link_qty is None:
            link_qty = 1.0
        return [(float(link_qty), None)]

    stack: List[Tuple[str,str,str,float]] = []
    for link_idx, l in enumerate(children(root_pn, root_rev), start=1):
        child_rev = _clean_rev(getattr(l, "child_rev", "") or "")
        for _occ_idx, (oqty, seq) in enumerate(_link_occurrences(l), start=1):
            path = _append_path("+", seq, link_idx)
            stack.append((l.child_pn, child_rev, path, oqty))

    while stack:
        pn, rev, level_path, q = stack.pop()
        key = (pn, _clean_rev(rev))
        occ.setdefault(key, []).append((level_path, q))
        if not include_consumed:
            pdoc = _part_by(pn, rev)
            procs = set(_part_processes(pdoc))
            if terminals and (procs & terminals):
                # stop at consumed parts
                continue
        for link_idx, l in enumerate(children(pn, rev), start=1):
            child_rev = _clean_rev(getattr(l, "child_rev", "") or "")
            for _occ_idx, (oqty, seq) in enumerate(_link_occurrences(l), start=1):
                cq = oqty * q
                child_path = _append_path(level_path, seq, link_idx)
                stack.append((l.child_pn, child_rev, child_path, cq))
    return occ


def _excel_bom_bytes(
    root_pn: str,
    root_rev: Optional[str],
    flat: List[Tuple[str,str,float]],
    occ: Dict[Tuple[str,str], List[Tuple[str,float]]],
    full_qty_map: Optional[Dict[Tuple[str,str], float]] = None,
    include_all_fields: bool = False,
    field_ids: Optional[List[str]] = None,
) -> bytes:
    def _sorted_occ(key: Tuple[str, str]):
        rows = list(occ.get(key, []))
        if not rows:
            return rows
        def _key(item):
            level = item[0]
            if isinstance(level, (int, float)):
                return (0, float(level))
            return (1, str(level))
        return sorted(rows, key=_key)

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
        from openpyxl.styles import Alignment, Font, PatternFill
        import json as _json
    except Exception:
        # Fallback: CSV in memory
        import csv
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["Part Number","Revision","Qty","Levels","Level Qtys"]) 
        for pn, rev, qty in flat:
            occ_list = _sorted_occ((pn, _norm_rev(rev)))
            lvls = ", ".join(str(l) for l,_ in occ_list)
            lq   = ", ".join(f"{q:g}" for _,q in occ_list)
            w.writerow([pn, rev, qty, lvls, lq])
        return buf.getvalue().encode("utf-8")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"BOM_{(root_pn or 'root')[:20]}"

    # Determine union of attribute keys (canonicalized using ALIASES on existing keys only)
    attr_keys: Set[str] = set()
    parts_cache: Dict[Tuple[str,str], Optional[Part]] = {}
    coverage_map: Dict[Tuple[str, str], Set[str]] = {}
    for pn, rev, _ in flat:
        key = (pn, _norm_rev(rev))
        if key not in parts_cache:
            parts_cache[key] = _part_by(pn, rev)
        p = parts_cache[key]
        if p and isinstance(getattr(p, 'attrs', None), dict):
            for k in (p.attrs or {}).keys():
                if not k:
                    continue
                kl = str(k).strip().lower()
                canon = ALIASES.get(kl, kl)
                if canon:
                    attr_keys.add(str(canon))
    unique_pns = list({pn for pn, _rev, _qty in flat if pn})
    if unique_pns:
        for row in PartFile.objects(part_number__in=unique_pns).only("part_number", "revision", "ext_group"):
            key = (str(row.part_number or ""), _norm_rev(getattr(row, "revision", "") or ""))
            coverage_map.setdefault(key, set()).add(str(getattr(row, "ext_group", "") or "").lower())
    field_config = get_field_config()
    excel_context = field_config.get("contexts", {}).get("excel_bom", {})
    field_defs = {field["id"]: field for field in field_config.get("fields", [])}
    valid_excel_ids = set(excel_context.get("allowed_field_ids") or [])
    selected_field_ids = [field_id for field_id in (field_ids or excel_context.get("default_field_ids") or []) if field_id in valid_excel_ids]
    for req in excel_context.get("required_field_ids") or []:
        if req in valid_excel_ids and req not in selected_field_ids:
            selected_field_ids.append(req)
    if not selected_field_ids:
        selected_field_ids = ["thumbnail", "part_number", "revision", "description", "total_qty", "level", "level_qty"]

    ignore = {field_id.lower() for field_id in selected_field_ids}
    ignore.update({"oem_partnumber", "approvedby", "approved_by", "approved", "approveddate", "approved_date"})
    header_attrs = sorted([k for k in attr_keys if k and k.lower() not in ignore]) if include_all_fields else []
    header_fields = list(selected_field_ids) + header_attrs
    header_labels = [
        str(field_defs.get(field_id, {}).get("label") or field_id.replace("_", " ").title())
        for field_id in selected_field_ids
    ] + header_attrs
    ws.append(header_labels)
    ws.freeze_panes = 'A2'

    # Column widths
    widths = {
        'thumbnail': 14,
        'part_number': 26,
        'revision': 10,
        'level': 10,
        'level_qty': 14,
        'description': 40,
        'total_qty': 14,
        'material': 18,
        'process': 18,
        'finish': 16,
        'mass': 12,
        'link': 30,
        'oem': 18,
        'oem_partnumber': 22,
    }
    for i, field_id in enumerate(header_fields, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(field_id, 16)

    # Map header name -> column index for robust addressing
    header_index: Dict[str, int] = { field_id: idx+1 for idx, field_id in enumerate(header_fields) }

    # Helper to find preview PNG path
    file_root = (current_app.config.get('FILE_ROOT_LOCAL') or '').rstrip('/\\')
    fallback_logo = _static_image_path('tinylogo.png')
    def preview_png_path(pn: str, rev: str) -> Optional[str]:
        q = PartFile.objects(part_number__iexact=pn, revision__iexact=_norm_rev(rev), ext_group='png', is_dwg=False).order_by('-mtime_iso')
        pf = q.first()
        if pf:
            thumb_rel = getattr(pf, "thumb_rel_path", None)
            if thumb_rel:
                thumb_abs = os.path.join(file_root, thumb_rel.replace("/", os.sep))
                if os.path.isfile(thumb_abs):
                    return thumb_abs
            pth = pf.path if os.path.isabs(pf.path) else os.path.join(file_root, pf.rel_path.replace('/', os.sep))
            if os.path.isfile(pth):
                return pth
        return fallback_logo

    # helper to coerce any value into an Excel-friendly scalar
    def _cell(v):
        if v is None:
            return ""
        if isinstance(v, (list, tuple, set)):
            return ", ".join(str(x) for x in v)
        # allow datetime/date as-is
        try:
            import datetime as __dt
            if isinstance(v, (__dt.date, __dt.datetime)):
                return v
        except Exception:
            pass
        if isinstance(v, dict):
            try:
                return _json.dumps(v, ensure_ascii=False)
            except Exception:
                return str(v)
        # keep numbers/bools
        if isinstance(v, (int, float, bool)):
            return v
        return str(v)

    # Write rows
    row_idx = 1
    for pn, rev, qty in flat:
        row_idx += 1
        key = (pn, _norm_rev(rev))
        p = parts_cache.get(key)
        # Normalize attributes to handle mixed-case and aliases (e.g., Link/LINK/oem_link -> link)
        attrs = harvest_part_attrs(p) if p else {}
        occ_list = _sorted_occ(key)
        levels_list = [l for l,_ in occ_list]
        qtys_list = [q for _,q in occ_list]
        full_qty = (full_qty_map or {}).get(key, qty)
        values = resolve_part_field_values(
            p,
            selected_field_ids,
            attrs=attrs,
            config=field_config,
            extra={
                "part_number": pn,
                "revision": _norm_rev(rev),
                "description": _part_description(p, attrs) if p else attrs.get('description','') or '',
                "total_qty": full_qty,
                "level": ", ".join(str(x) for x in levels_list),
                "level_qty": ", ".join(f"{x:g}" for x in qtys_list),
            },
            coverage=coverage_map.get(key) or set(),
        )
        # Merge all attributes at the end, blanks when missing
        for k in header_attrs:
            values[k] = attrs.get(k, '')
        row = [_cell(values.get(k, '')) for k in header_fields]
        ws.append(row)
        # Adjust row height for thumbnail
        ws.row_dimensions[row_idx].height = 56
        # Add image if we have a path
        img_path = preview_png_path(pn, _norm_rev(rev))
        if img_path and os.path.isfile(img_path):
            try:
                img = XLImage(img_path)
                img.height = 48
                img.width = 48
                # Anchor to the 'thumbnail' cell (A{row}) using TwoCellAnchor so it moves/sizes with cell
                # openpyxl uses 0-based indices for AnchorMarker
                r0 = row_idx - 1
                img.anchor = TwoCellAnchor(
                    _from=AnchorMarker(col=0, colOff=5_000, row=r0, rowOff=5_000),
                    to=AnchorMarker(col=0, colOff=500_000, row=r0+1, rowOff=0),
                    editAs='twoCell'
                )
                ws.add_image(img)
            except Exception:
                pass
        # Hyperlink partnumber to the app part-detail page
        try:
            app_url = _part_detail_url(pn, _norm_rev(rev))
            pn_col = header_index.get('part_number')
            if pn_col:
                pcell = ws.cell(row=row_idx, column=pn_col)
                pcell.hyperlink = app_url
                pcell.font = Font(color='0000EE', underline='single')
        except Exception:
            pass

        # Hyperlink on link column if URL present (external only)
        try:
            raw = str(values.get('link') or '').strip()
            if raw:
                url = raw
                if not (url.startswith('http://') or url.startswith('https://')):
                    url = 'http://' + url
                lcol = header_index.get('link')
                if lcol:
                    lcell = ws.cell(row=row_idx, column=lcol)
                    # HYPERLINK formula for compatibility
                    disp = raw
                    lcell.value = f'=HYPERLINK("{url}","{disp}")'
                try:
                    lcell.hyperlink = url
                except Exception:
                    pass
                lcell.font = Font(color='0000EE', underline='single')
        except Exception:
            pass

    # Header style & autofilter
    bold = Font(b=True)
    center = Alignment(horizontal='center')
    main_fill = PatternFill(fill_type='solid', fgColor='000000')
    main_font = Font(b=True, color='FFFFFF')
    for col_idx in range(1, len(header_fields)+1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold
        cell.alignment = center
        if col_idx <= len(selected_field_ids):
            cell.fill = main_fill
            cell.font = main_font
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header_fields))}{row_idx}"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _part_detail_url(pn: str, rev: str) -> str:
    rev_val = _clean_rev(rev)
    try:
        base = url_for("ui.part_ui", pn=pn, _external=True)
    except Exception:
        try:
            root = (request.url_root or "").rstrip("/")
        except Exception:
            root = ""
        if not root:
            root = (current_app.config.get("VITE_BACKEND_URL") or "http://localhost:5000").rstrip("/")
        pn_enc = quote(str(pn or ""), safe="")
        base = f"{root}/ui/part/{pn_enc}"
    if rev_val:
        rev_enc = quote(str(rev_val), safe="")
        sep = "&" if "?" in base else "?"
        return f"{base}{sep}rev={rev_enc}"
    return base


def _pil_to_rl_image(pil_img, width=None, height=None):
    from reportlab.lib.utils import ImageReader
    return ImageReader(pil_img)


def _process_color_map() -> Dict[str, Tuple[float,float,float]]:
    meta = current_app.config.get("PROCESS_META", {}) or {}
    out: Dict[str, Tuple[float,float,float]] = {}
    for k, v in meta.items():
        if k.startswith("_"): continue
        try:
            r, g, b = [int(x.strip()) for x in (v.get("color") or "").split(",")[:3]]
            out[k.lower()] = (max(0,min(r,255))/255.0, max(0,min(g,255))/255.0, max(0,min(b,255))/255.0)
        except Exception:
            pass
    return out


_BRAND_LOGO_NAMES = {
    "logo.svg",
    "logo.png",
    "tinylogo.svg",
    "tinylogo.png",
    "tinyfooter.svg",
    "tinyfooter.png",
}


def _branding_logo_paths() -> Tuple[Optional[str], Optional[str]]:
    path = resolve_brand_logo_path()
    if not path:
        return (None, None)
    ext = os.path.splitext(path)[1].lower()
    if ext == ".svg":
        return (path, None)
    if ext == ".png":
        return (None, path)
    return (None, None)


def _static_image_path(*names: str) -> Optional[str]:
    brand_svg, brand_png = _branding_logo_paths()
    base = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'static', 'images'))
    for n in names:
        nl = (n or "").lower()
        if nl in _BRAND_LOGO_NAMES:
            if nl.endswith(".svg") and brand_svg:
                return brand_svg
            if nl.endswith(".png") and brand_png:
                return brand_png
        p = os.path.join(base, n)
        if os.path.isfile(p):
            return p
    return None


def _draw_svg_or_png(c, x, y, w, h, svg_name: str, png_fallback: str):
    prefer_brand = False
    for nm in (svg_name, png_fallback):
        if (nm or "").lower() in _BRAND_LOGO_NAMES:
            prefer_brand = True
            break
    brand_svg, brand_png = _branding_logo_paths()
    if prefer_brand and brand_svg:
        try:
            from svglib.svglib import svg2rlg
            from reportlab.graphics import renderPDF
            drawing = svg2rlg(brand_svg)
            sx = w / float(drawing.width or 1)
            sy = h / float(drawing.height or 1)
            s = min(sx, sy)
            c.saveState(); c.translate(x, y); c.scale(s, s)
            renderPDF.draw(drawing, c, 0, 0)
            c.restoreState(); return
        except Exception:
            pass
    if prefer_brand and brand_png:
        try:
            c.drawImage(brand_png, x, y, width=w, height=h, preserveAspectRatio=True, anchor='sw', mask='auto')
            return
        except Exception:
            pass
    svg_path = _static_image_path(svg_name)
    if svg_path:
        try:
            from svglib.svglib import svg2rlg
            from reportlab.graphics import renderPDF
            drawing = svg2rlg(svg_path)
            sx = w / float(drawing.width or 1)
            sy = h / float(drawing.height or 1)
            s = min(sx, sy)
            c.saveState(); c.translate(x, y); c.scale(s, s)
            renderPDF.draw(drawing, c, 0, 0)
            c.restoreState(); return
        except Exception:
            pass
    png = _static_image_path(png_fallback)
    if png:
        try:
            c.drawImage(png, x, y, width=w, height=h, preserveAspectRatio=True, anchor='sw', mask='auto')
        except Exception:
            pass


def _visual_list_pdf(
    flat: List[Tuple[str,str,float]],
    root_pn: Optional[str] = None,
    root_rev: Optional[str] = None,
    *,
    page_map: Optional[Dict[Tuple[str,str], int]] = None,
    build_ts: Optional[datetime] = None,
    hardware_rows: Optional[List[Dict[str, object]]] = None,
    hardware_pdf_bytes: Optional[bytes] = None,
    root_desc: Optional[str] = None,
    root_thumb_override: Optional[str] = None,
) -> Optional[bytes]:
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfbase.pdfmetrics import stringWidth
        import qrcode
        from PIL import Image as PILImage
    except Exception:
        return None
    root = (current_app.config.get("FILE_ROOT_LOCAL") or "").rstrip("/\\")
    root_rev_clean = _clean_rev(root_rev) if root_rev is not None else ""
    resolved_root_desc = root_desc
    if resolved_root_desc is None and root_pn:
        try:
            rpdoc = _part_by(root_pn, root_rev_clean)
            resolved_root_desc = _part_description(rpdoc)
        except Exception:
            resolved_root_desc = ""
    if resolved_root_desc is None:
        resolved_root_desc = ""

    def _clip_line(text: str, font: str, size: float, max_w: float) -> str:
        text = text or ""
        if not text:
            return ""
        if stringWidth(text, font, size) <= max_w:
            return text
        suffix = "..."
        max_w = max(0.0, max_w - stringWidth(suffix, font, size))
        out = text
        while out and stringWidth(out, font, size) > max_w:
            out = out[:-1]
        return out.rstrip() + (suffix if out else "")
    def _wrap_lines(text: str, font: str, size: float, max_w: float, max_lines: int) -> List[str]:
        text = (text or "").strip()
        if not text or max_lines <= 0:
            return []
        words = text.split()
        lines: List[str] = []
        cur = ""
        truncated = False
        for w in words:
            cand = (cur + " " + w).strip()
            if stringWidth(cand, font, size) <= max_w:
                cur = cand
            else:
                if cur:
                    lines.append(cur)
                cur = w
                if len(lines) >= max_lines:
                    truncated = True
                    cur = ""
                    break
        if cur and len(lines) < max_lines:
            lines.append(cur)
        if truncated and lines:
            lines[-1] = _clip_line(lines[-1], font, size, max_w)
        return lines

    # Build rows with: pn, rev, qty, imgpath
    # Ensure root appears first (alone) and children sorted by partnumber
    rows: List[Tuple[str,str,float,str|None]] = []
    # Prepare children items, excluding the root if present
    items = list(flat or [])
    if root_pn:
        items = [t for t in items if not (str(t[0]).strip().lower() == str(root_pn).strip().lower() and _norm_rev(t[1]) == root_rev_clean)]
    # Sort children by partnumber then revision
    items.sort(key=lambda t: (str(t[0]) or "", _norm_rev(t[1]) or ""))

    # Helper to find image path for a pn/rev
    def _img_for(pn: str, rev: str) -> Optional[str]:
        q = PartFile.objects(part_number__iexact=pn, revision__iexact=_clean_rev(rev), ext_group="png", is_dwg=False).order_by("-mtime_iso")
        pf = q.first()
        if pf:
            thumb_rel = getattr(pf, "thumb_rel_path", None)
            if thumb_rel:
                thumb_abs = os.path.join(root, thumb_rel.replace("/", os.sep))
                if os.path.isfile(thumb_abs):
                    return thumb_abs
            pth = pf.path if os.path.isabs(pf.path) else os.path.join(root, pf.rel_path.replace("/", os.sep))
            if os.path.isfile(pth):
                return pth
        return None

    # Root entry first if provided
    root_entry: Optional[Tuple[str,str,float,Optional[str]]] = None
    if root_pn:
        root_img = root_thumb_override if (root_thumb_override and os.path.isfile(root_thumb_override)) else _img_for(root_pn, root_rev_clean)
        root_entry = (root_pn, root_rev_clean, 1.0, root_img)

    # Child rows
    for pn, rev, qty in items:
        rows.append((pn, _clean_rev(rev), qty, _img_for(pn, _clean_rev(rev))))

    # Layout close to legacy BoxyGrid (3 columns on A4 portrait)
    box_w = 58*mm
    box_h = 46*mm
    gap = 6*mm
    margin = 14*mm
    header_pad = 10*mm
    W, H = A4
    cols = max(1, int((W - 2*margin + gap) // (box_w + gap)))
    # Aim for exactly 3 columns if space permits
    cols = min(max(cols, 3), 4)
    x0 = margin
    y0 = H - margin - header_pad
    qty_margin = 8*mm

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    # Title with root PN/REV and description, plus logos
    def header():
        title = (f"{root_pn or ''} REV {root_rev_clean}").strip()
        c.setFont("Helvetica-Bold", 18)
        if title:
            c.drawCentredString(W/2, H - margin + 6*mm, title)
        if resolved_root_desc:
            c.setFont("Helvetica", 11)
            c.drawCentredString(W/2, H - margin - 1*mm, _clip_line(resolved_root_desc[:200], "Helvetica", 11, W - 2*margin))
        stamp = format_display_ts(build_ts)
        if stamp:
            c.setFont("Helvetica", 9.5)
            c.setFillGray(0.45)
            c.drawCentredString(W/2, H - margin - 7*mm, f"Generated: {stamp}")
            c.setFillGray(0.1)
        # top-left logo (10mm)
        _draw_svg_or_png(c, margin-8*mm, H - margin + 0.5*mm, 10*mm, 10*mm, 'logo.svg', 'tinylogo.png')
        # bottom center footer

    header()

    proc_colors = _process_color_map()
    rows_per_page = max(1, int((H - 2*margin - header_pad + gap) // (box_h + gap)))
    cur_page = 0

    def draw_cell(ix: int, pn: str, rev: str, qty: float, imgpath: Optional[str], desc: str = ""):
        nonlocal cur_page
        page_no = ix // (cols * rows_per_page)
        if page_no != cur_page:
            c.showPage(); header(); cur_page = page_no
        pos = ix % (cols * rows_per_page)
        col = pos % cols
        row_in_page = pos // cols
        # compute origin of the cell (top-left)
        x = x0 + col * (box_w + gap)
        y = y0 - (row_in_page + 1) * (box_h + gap)
        # base box (darker stroke for consistent visibility)
        c.setStrokeColorRGB(0.1, 0.1, 0.1); c.setLineWidth(1.6)
        c.setFillColorRGB(1,1,1)
        c.rect(x, y, box_w, box_h, stroke=1, fill=1)
        # process colored rings inside the border
        pdoc = _part_by(pn, rev)
        procs = _part_processes(pdoc)
        procs_lower = [p.lower() for p in procs]
        inset = 0.8*mm
        for i, p in enumerate(procs_lower):
            col = proc_colors.get(p)
            if not col: continue
            off = (i+1) * inset
            c.setStrokeColorRGB(*col); c.setLineWidth(2)
            c.rect(x+off, y+off, box_w-2*off, box_h-2*off, stroke=1, fill=0)
        # geometry: reserve top header (PN/REV/QTY) and bottom (desc + QR)
        top_space = 16*mm
        bottom_space = 8*mm
        # image area (left side), vertically between top_space and bottom_space
        ix_x = x + 3*mm
        ix_y = y + bottom_space
        ix_w = 34*mm
        ix_h = box_h - (top_space + bottom_space)
        if imgpath:
            try:
                c.drawImage(imgpath, ix_x, ix_y, width=ix_w, height=ix_h, preserveAspectRatio=True, anchor='sw', mask='auto')
            except Exception:
                pass
        else:
            _draw_svg_or_png(c, ix_x, ix_y, ix_w, ix_h, 'tinylogo.svg', 'tinylogo.png')
        # text area
        hl_x = x + 4*mm
        hl_y = y + box_h - 6*mm
        if not desc and pdoc is not None:
            try:
                desc = _part_description(pdoc)
            except Exception:
                desc = ''
        try:
            qty_val = float(qty)
            qty_str = f"x{qty_val:g}"
        except Exception:
            qty_str = f"x{qty}"
        try:
            from reportlab.pdfbase.pdfmetrics import stringWidth
            qty_w = stringWidth(qty_str, "Helvetica-Bold", 11.5)
        except Exception:
            qty_w = 0.0
        title = f"{pn} REV {rev}" if rev else f"{pn}"
        max_title_w = max(20*mm, box_w - (hl_x - x) - qty_w - 6*mm)
        c.setFillGray(0.1)
        c.setFont("Helvetica-Bold", 11.5)
        c.drawString(hl_x, hl_y, _clip_line(title, "Helvetica-Bold", 11.5, max_title_w))
        qr_size = 7*mm
        icon_w = 8*mm
        icon_h = 6*mm
        desc_line_h = 0.0
        desc_lines_count = 0
        if desc:
            desc_font = 8.8
            desc_line_h = desc_font * 1.25
            icon_top = y + 4*mm + qr_size + 1*mm + icon_h
            desc_y = hl_y - 12
            desc_max_w = box_w - (hl_x - x) - (qr_size + 6*mm)
            available_h = max(0.0, desc_y - (icon_top + 2*mm))
            max_lines = max(1, int(available_h / desc_line_h) + 1)
            max_lines = min(max_lines, 4)
            lines = _wrap_lines(desc, "Helvetica", desc_font, desc_max_w, max_lines)
            desc_lines_count = len(lines)
            c.setFont("Helvetica", desc_font)
            c.setFillGray(0.35)
            for i, line in enumerate(lines):
                c.drawString(hl_x, desc_y - (i * desc_line_h), line)
            c.setFillGray(0.1)
        # Qty at top-right in bold red
        try:
            if qty_str:
                c.setFont("Helvetica-Bold", 11.5)
                c.setFillColorRGB(0.82, 0.0, 0.0)
                qty_x = x + box_w - qty_margin - qty_w
                c.drawString(qty_x, hl_y, qty_str)
                c.setFillGray(0.1)
                c.setFillColorRGB(0.1, 0.1, 0.1)
        except Exception:
            pass
        # Binder page reference (optional): show below quantity if provided
        try:
            if page_map is not None:
                key = (pn, _norm_rev(rev))
                pg = page_map.get(key)
                if pg is not None:
                    from reportlab.pdfbase.pdfmetrics import stringWidth
                    s = f"p. {pg}"
                    c.setFont("Helvetica", 8.5)
                    c.setFillGray(0.4)
                    tw = stringWidth(s, "Helvetica", 8.5)
                    icon_top = y + 4*mm + qr_size + 1*mm + icon_h
                    desc_y = hl_y - 12
                    lines_used = desc_lines_count or 1
                    line_h = desc_line_h or (8.8 * 1.25)
                    page_y = max(icon_top + 1*mm, desc_y - (line_h * (lines_used + 1)))
                    c.drawString(x + box_w - 4*mm - tw, page_y, s)
                    c.setFillGray(0.1)
        except Exception:
            pass
        # QR code (links to part detail) — draw last to ensure it sits on top
        try:
            rev_qr = rev
            if (not rev_qr) and pdoc is not None:
                rev_qr = _clean_rev(getattr(pdoc, "revision", "") or "")
            qr_url = _part_detail_url(pn, rev_qr)
            qr = qrcode.QRCode(border=0, box_size=2)
            qr.add_data(qr_url)
            qr.make(fit=True)
            qimg = qr.make_image(fill_color="black", back_color="white").convert("RGB")
            from reportlab.lib.utils import ImageReader
            qbuf = io.BytesIO(); qimg.save(qbuf, format='PNG'); qbuf.seek(0)
            c.drawImage(ImageReader(qbuf), x + box_w - qr_size - 4*mm, y + 4*mm, width=qr_size, height=qr_size, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass

        # Approval status icon (above the QR)
        try:
            approved = False
            if pdoc is not None:
                a = harvest_part_attrs(pdoc)
                approved = bool(approved_value(a))
            show_notapproved = True
            if procs_lower and any(p in ('hardware', 'purchase', 'others') for p in procs_lower):
                show_notapproved = False
            icon_x = x + box_w - icon_w - 4*mm
            icon_y = y + 4*mm + qr_size + 1*mm
            if approved:
                _draw_svg_or_png(c, icon_x, icon_y, icon_w, icon_h, 'approved.svg', 'approved.png')
            elif show_notapproved:
                _draw_svg_or_png(c, icon_x, icon_y, icon_w, icon_h, 'notapproved.svg', 'notapproved.png')
        except Exception:
            pass

    # Special root cell: render alone on the first row (spanning all columns)
    if root_entry is not None:
        try:
            pn, rev, qty, ip = root_entry
            # compute a full-width box occupying the first grid row height
            r_x = x0
            r_y = y0 - (box_h + gap)
            r_w = min(W - 2*margin, cols * (box_w + gap) - gap)
            r_h = box_h
            # base box (darker stroke for consistent visibility)
            c.setStrokeColorRGB(0.1, 0.1, 0.1); c.setLineWidth(1.6)
            c.setFillColorRGB(1,1,1)
            c.rect(r_x, r_y, r_w, r_h, stroke=1, fill=1)
            # image area on left, wider than normal
            ix_x = r_x + 3*mm
            ix_y = r_y + 8*mm
            ix_w = 40*mm
            ix_h = r_h - 16*mm
            if ip and os.path.isfile(ip):
                try:
                    c.drawImage(ip, ix_x, ix_y, width=ix_w, height=ix_h, preserveAspectRatio=True, anchor='sw', mask='auto')
                except Exception:
                    _draw_svg_or_png(c, ix_x, ix_y, ix_w, ix_h, 'tinylogo.svg', 'tinylogo.png')
            else:
                _draw_svg_or_png(c, ix_x, ix_y, ix_w, ix_h, 'tinylogo.svg', 'tinylogo.png')
            # text header
            hl_x = r_x + ix_w + 6*mm
            hl_y = r_y + r_h - 6*mm
            try:
                pdoc = _part_by(pn, rev)
                desc = _part_description(pdoc)
            except Exception:
                desc = ''
            try:
                qty_val = float(qty)
                qty_str = f"x{qty_val:g}"
            except Exception:
                qty_str = f"x{qty}"
            try:
                qty_w = stringWidth(qty_str, "Helvetica-Bold", 12.5)
            except Exception:
                qty_w = 0.0
            title = f"{pn} REV {rev}" if rev else f"{pn}"
            max_title_w = max(25*mm, r_w - (hl_x - r_x) - qty_w - 6*mm)
            c.setFillGray(0.1)
            c.setFont("Helvetica-Bold", 12.5)
            c.drawString(hl_x, hl_y, _clip_line(title, "Helvetica-Bold", 12.5, max_title_w))
            desc_line_h = 0.0
            desc_lines_count = 0
            if desc:
                desc_font = 9.0
                desc_line_h = desc_font * 1.25
                desc_y = hl_y - 14
                desc_max_w = r_w - (hl_x - r_x) - 20*mm
                available_h = max(0.0, desc_y - (r_y + 6*mm + 10*mm))
                max_lines = max(1, int(available_h / desc_line_h) + 1)
                max_lines = min(max_lines, 4)
                lines = _wrap_lines(desc, "Helvetica", desc_font, desc_max_w, max_lines)
                desc_lines_count = len(lines)
                c.setFont("Helvetica", desc_font)
                c.setFillGray(0.35)
                for i, line in enumerate(lines):
                    c.drawString(hl_x, desc_y - (i * desc_line_h), line)
                c.setFillGray(0.1)
            # qty at top-right in bold red
            try:
                c.setFont("Helvetica-Bold", 12.5)
                c.setFillColorRGB(0.82, 0.0, 0.0)
                tw = stringWidth(qty_str, "Helvetica-Bold", 12.5)
                c.drawString(r_x + r_w - qty_margin - tw, hl_y, qty_str)
                c.setFillGray(0.1)
                c.setFillColorRGB(0.1, 0.1, 0.1)
                # binder page below description if available
                if page_map is not None:
                    key = (pn, _norm_rev(rev))
                    pg = page_map.get(key)
                    if pg is not None:
                        s = f"p. {pg}"
                        c.setFont("Helvetica", 9.0); c.setFillGray(0.4)
                        tw2 = stringWidth(s, "Helvetica", 9.0)
                        lines_used = desc_lines_count or 1
                        line_h = desc_line_h or (9.0 * 1.25)
                        page_y = hl_y - 14 - (line_h * (lines_used + 1))
                        c.drawString(r_x + r_w - qty_margin - tw2, page_y, s)
                        c.setFillGray(0.1)
            except Exception:
                pass
            # QR bottom-right
            try:
                rev_qr = rev
                if (not rev_qr) and pdoc is not None:
                    rev_qr = _clean_rev(getattr(pdoc, "revision", "") or "")
                qr_url = _part_detail_url(pn, rev_qr)
                qr = qrcode.QRCode(border=0, box_size=2)
                qr.add_data(qr_url)
                qr.make(fit=True)
                qimg = qr.make_image(fill_color="black", back_color="white").convert("RGB")
                from reportlab.lib.utils import ImageReader
                qbuf = io.BytesIO(); qimg.save(qbuf, format='PNG'); qbuf.seek(0)
                qr_size = 10*mm
                c.drawImage(ImageReader(qbuf), r_x + r_w - qr_size - 4*mm, r_y + 4*mm, width=qr_size, height=qr_size, preserveAspectRatio=True, mask='auto')
            except Exception:
                pass
        except Exception:
            pass

    # Draw children; offset grid index to skip first row reserved by root
    start_ix = cols if root_entry is not None else 0
    for i, (pn, rev, qty, ip) in enumerate(rows):
        try:
            pdoc = _part_by(pn, rev)
            desc = _part_description(pdoc)
        except Exception:
            desc = ''
        draw_cell(start_ix + i, pn, rev, qty, ip, desc)

    c.save()
    base_pdf = buf.getvalue()
    if hardware_rows:
        try:
            hw_pdf = _hardware_summary_pdf_with_empty(
                hardware_rows,
                allow_empty=False,
                root_pn=root_pn,
                root_rev=root_rev,
                root_desc=resolved_root_desc,
                build_ts=build_ts,
            )
        except Exception:
            hw_pdf = None
        if hw_pdf:
            try:
                from PyPDF2 import PdfReader, PdfWriter
            except Exception:
                return base_pdf
            writer = PdfWriter()
            for payload in (base_pdf, hw_pdf):
                reader = PdfReader(io.BytesIO(payload))
                for page in reader.pages:
                    writer.add_page(page)
            merged_buf = io.BytesIO()
            writer.write(merged_buf)
            return merged_buf.getvalue()
    return base_pdf


def _norm_proc_name(val: object) -> str:
    return " ".join(str(val or "").strip().lower().split())


def _is_hardware_processes(processes: Iterable[str], alias_index: Dict[str, str]) -> bool:
    for p in processes:
        key = _norm_proc_name(p)
        canon = alias_index.get(key, key)
        if canon == "hardware":
            return True
    return False


def _category_override_is_non_hardware(pdoc: Part, attrs: Dict[str, object]) -> bool:
    cat = str(getattr(pdoc, "category", "") or attrs.get("category") or "").strip().lower()
    return bool(cat and "hardware" not in cat)


def _is_hardware_part(pn: str, rev: str) -> bool:
    pdoc = _part_by(pn, rev)
    if not pdoc:
        return False
    attrs = harvest_part_attrs(pdoc)
    if _category_override_is_non_hardware(pdoc, attrs):
        return False
    meta = current_app.config.get("PROCESS_META", {}) or {}
    classification = classify_part(attrs, list(pdoc.processes or []), meta, category=getattr(pdoc, "category", ""))
    return classification == "hardware"


def _hardware_summary_rows(pn_rev_qty: Iterable[Tuple[str, str, float]]) -> List[Dict[str, object]]:
    rows: Dict[Tuple[str, str, str, str, str], float] = {}
    for pn, rev, qty in pn_rev_qty:
        pdoc = _part_by(pn, rev)
        if not pdoc:
            continue
        attrs = harvest_part_attrs(pdoc)
        if _category_override_is_non_hardware(pdoc, attrs):
            continue
        meta = current_app.config.get("PROCESS_META", {}) or {}
        classification = classify_part(attrs, list(pdoc.processes or []), meta, category=getattr(pdoc, "category", ""))
        if classification != "hardware":
            continue
        resolved = resolve_part_field_values(
            pdoc,
            ["material", "finish"],
            attrs=attrs,
            config=get_field_config(),
            extra={"part_number": pn, "revision": _norm_rev(rev)},
        )
        desc = _part_description(pdoc, attrs)
        material = resolved.get("material", "") or ""
        finish = resolved.get("finish", "") or ""
        key = (pn, _norm_rev(rev), desc, material, finish)
        rows[key] = rows.get(key, 0.0) + float(qty or 0.0)
    out = []
    for (pn, rev, desc, material, finish), qty in rows.items():
        out.append({
            "partnumber": pn,
            "revision": rev,
            "description": desc,
            "material": material,
            "finish": finish,
            "qty": qty,
        })
    out.sort(key=lambda r: (str(r.get("partnumber") or ""), str(r.get("revision") or "")))
    return out


def _hardware_summary_pdf(
    rows: List[Dict[str, object]],
    root_pn: Optional[str] = None,
    root_rev: Optional[str] = None,
    root_desc: Optional[str] = None,
    build_ts: Optional[datetime] = None,
) -> Optional[bytes]:
    return _hardware_summary_pdf_with_empty(
        rows,
        allow_empty=False,
        root_pn=root_pn,
        root_rev=root_rev,
        root_desc=root_desc,
        build_ts=build_ts,
    )


def _hardware_summary_pdf_with_empty(
    rows: List[Dict[str, object]],
    allow_empty: bool = False,
    root_pn: Optional[str] = None,
    root_rev: Optional[str] = None,
    root_desc: Optional[str] = None,
    build_ts: Optional[datetime] = None,
) -> Optional[bytes]:
    if not rows and not allow_empty:
        return None
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from xml.sax.saxutils import escape as _xml_escape
    except Exception:
        return None

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle(
        "hw_cell",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8,
        leading=9,
        wordWrap="CJK",
    )
    header_style = ParagraphStyle(
        "hw_header",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=8.5,
        leading=9.5,
        alignment=1,
    )
    qty_style = ParagraphStyle(
        "hw_qty",
        parent=cell_style,
        alignment=2,
    )
    flowables = []
    flowables.append(Paragraph("Hardware Summary", styles["Heading2"]))
    flowables.append(Paragraph("Hardware Summary Details", styles["Normal"]))
    if root_pn:
        root_rev_clean = _clean_rev(root_rev) if root_rev is not None else ""
        root_line = f"{root_pn} REV {root_rev_clean}".strip()
        flowables.append(Paragraph(_xml_escape(root_line), styles["Heading4"]))
        if root_desc:
            flowables.append(Paragraph(_xml_escape(str(root_desc)), cell_style))
        stamp = format_display_ts(build_ts)
        if stamp:
            flowables.append(Paragraph(_xml_escape(f"Generated: {stamp}"), cell_style))
    flowables.append(Spacer(0, 6*mm))
    if not rows:
        flowables.append(Paragraph("No hardware items found.", styles["Normal"]))
        doc.build(flowables)
        return buf.getvalue()

    table_data = [[
        Paragraph("No.", header_style),
        Paragraph("Part Number", header_style),
        Paragraph("Description", header_style),
        Paragraph("Material", header_style),
        Paragraph("Qty", header_style),
    ]]
    for idx, r in enumerate(rows, start=1):
        pn = str(r.get("partnumber") or "")
        desc = str(r.get("description") or "")
        material = str(r.get("material") or "")
        qty = r.get("qty") or 0
        try:
            qty_s = f"{float(qty):g}"
        except Exception:
            qty_s = str(qty)
        table_data.append([
            Paragraph(_xml_escape(str(idx)), qty_style),
            Paragraph(_xml_escape(pn), cell_style),
            Paragraph(_xml_escape(desc), cell_style),
            Paragraph(_xml_escape(material), cell_style),
            Paragraph(_xml_escape(qty_s), qty_style),
        ])

    usable_w = doc.width
    col_weights = [0.06, 0.2, 0.5, 0.16, 0.08]
    col_widths = [usable_w * w for w in col_weights]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (1, 1), (-2, -1), 'LEFT'),
        ('ALIGN', (-1, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8.5),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
    ])
    table.setStyle(style)
    flowables.append(table)
    doc.build(flowables)
    return buf.getvalue()

def _aggregate_qty(pn_rev_qty: Iterable[Tuple[str, str, float]]) -> Dict[Tuple[str, str], float]:
    totals: Dict[Tuple[str, str], float] = {}
    for pn, rev, qty in pn_rev_qty:
        key = (pn, _clean_rev(rev))
        totals[key] = totals.get(key, 0.0) + float(qty or 0.0)
    return totals

def _scope_supply_rows(pn_rev_qty: Iterable[Tuple[str, str, float]]) -> List[Dict[str, object]]:
    totals = _aggregate_qty(pn_rev_qty)
    rows: List[Dict[str, object]] = []
    for (pn, rev), qty in sorted(totals.items(), key=lambda t: (t[0][0] or "", t[0][1] or "")):
        pdoc = _part_by(pn, rev)
        rows.append(
            {
                "partnumber": pn,
                "revision": rev,
                "description": _part_description(pdoc),
                "qty": qty,
            }
        )
    return rows

def _cut_fold_summary_rows(pn_rev_qty: Iterable[Tuple[str, str, float]]) -> List[Dict[str, object]]:
    totals = _aggregate_qty(pn_rev_qty)
    rows: List[Dict[str, object]] = []
    for (pn, rev), qty in sorted(totals.items(), key=lambda t: (t[0][0] or "", t[0][1] or "")):
        pdoc = _part_by(pn, rev)
        procs = _part_processes(pdoc)
        rows.append(
            {
                "partnumber": pn,
                "revision": rev,
                "description": _part_description(pdoc),
                "process": ", ".join(procs),
                "qty": qty,
            }
        )
    return rows

def _scope_supply_pdf(rows: List[Dict[str, object]], title: str) -> Optional[bytes]:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet
    except Exception:
        return None

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    flowables = [Paragraph(title, styles["Heading2"]), Spacer(0, 4*mm)]
    if not rows:
        flowables.append(Paragraph("No items found.", styles["Normal"]))
        doc.build(flowables)
        return buf.getvalue()

    table_data = [["Part Number", "Revision", "Description", "Total Qty"]]
    for r in rows:
        qty = r.get("qty") or 0
        try:
            qty_s = f"{float(qty):g}"
        except Exception:
            qty_s = str(qty)
        table_data.append([
            str(r.get("partnumber") or ""),
            str(r.get("revision") or ""),
            str(r.get("description") or ""),
            qty_s,
        ])

    col_widths = [32*mm, 16*mm, 90*mm, 20*mm]
    table = Table(table_data, colWidths=col_widths)
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
    ])
    table.setStyle(style)
    flowables.append(table)
    doc.build(flowables)
    return buf.getvalue()

def _cut_fold_summary_pdf(rows: List[Dict[str, object]], title: str) -> Optional[bytes]:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet
    except Exception:
        return None

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    flowables = [Paragraph(title, styles["Heading2"]), Spacer(0, 4*mm)]
    if not rows:
        flowables.append(Paragraph("No items found.", styles["Normal"]))
        doc.build(flowables)
        return buf.getvalue()

    table_data = [["Part Number", "Revision", "Description", "Process", "Total Qty"]]
    for r in rows:
        qty = r.get("qty") or 0
        try:
            qty_s = f"{float(qty):g}"
        except Exception:
            qty_s = str(qty)
        table_data.append([
            str(r.get("partnumber") or ""),
            str(r.get("revision") or ""),
            str(r.get("description") or ""),
            str(r.get("process") or ""),
            qty_s,
        ])

    col_widths = [30*mm, 14*mm, 70*mm, 45*mm, 18*mm]
    table = Table(table_data, colWidths=col_widths)
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
    ])
    table.setStyle(style)
    flowables.append(table)
    doc.build(flowables)
    return buf.getvalue()


def _whereused_rows(root_pn: str, root_rev: Optional[str]) -> List[Dict[str, object]]:
    rows: Dict[Tuple[str, str], Dict[str, object]] = {}
    if not root_pn:
        return []
    try:
        if "child_pn" in BOMLink._fields:
            q = BOMLink.objects(child_pn=root_pn)
            root_rev_clean = _rev_or_none(root_rev)
            if root_rev_clean is not None and "child_rev" in BOMLink._fields:
                q = q.filter(child_rev=root_rev_clean)
            links = q
        else:
            child_part = Part.objects(part_number=root_pn).only("id").first()
            links = BOMLink.objects(child=child_part) if child_part else []
    except Exception:
        links = []

    for l in links:
        if "parent_pn" in BOMLink._fields:
            parent_pn = getattr(l, "parent_pn", None)
            parent_rev = _clean_rev(getattr(l, "parent_rev", "") if hasattr(l, "parent_rev") else "")
        else:
            parent_obj = getattr(l, "parent", None)
            parent_pn = getattr(parent_obj, "part_number", None)
            parent_rev = _clean_rev(getattr(parent_obj, "revision", "") if parent_obj else "")
        if not parent_pn:
            continue
        key = (parent_pn, parent_rev or "")
        row = rows.get(key)
        if row is None:
            pdoc = _part_by(parent_pn, parent_rev or "")
            desc = _part_description(pdoc) if pdoc else ""
            row = {
                "parent_pn": parent_pn,
                "parent_rev": parent_rev or "",
                "description": desc,
                "qty": 0.0,
            }
            rows[key] = row
        qty = getattr(l, "qty", 0) or 0
        try:
            row["qty"] = float(row.get("qty") or 0.0) + float(qty)
        except Exception:
            pass

    return sorted(rows.values(), key=lambda r: (r.get("parent_pn") or "", r.get("parent_rev") or ""))


def _whereused_report_pdf(
    rows: List[Dict[str, object]],
    root_pn: str,
    root_rev: Optional[str],
    build_ts: Optional[datetime] = None,
) -> Optional[bytes]:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
        from reportlab.lib.styles import getSampleStyleSheet
    except Exception:
        return None

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()

    flowables = []
    flowables.append(Paragraph("Where Used", styles["Heading2"]))
    flowables.append(Spacer(0, 4*mm))

    ts = format_display_ts(build_ts)
    if root_pn:
        rev_s = _clean_rev(root_rev) if root_rev is not None else ""
        flowables.append(Paragraph(f"Part: {root_pn}    Rev: {rev_s}", styles["Normal"]))
    if ts:
        flowables.append(Paragraph(f"Generated: {ts}", styles["Normal"]))
    flowables.append(Spacer(0, 6*mm))

    if not rows:
        flowables.append(Paragraph("No where-used records found.", styles["Normal"]))
        doc.build(flowables)
        return buf.getvalue()

    file_root = (current_app.config.get("FILE_ROOT_LOCAL") or current_app.config.get("FILES_LOCAL_ROOT") or "").rstrip("/\\")
    def preview_png_path(pn: str, rev: str) -> Optional[str]:
        try:
            q = PartFile.objects(part_number__iexact=pn, ext_group="png", is_dwg=False)
            q = q.filter(revision__iexact=_norm_rev(rev))
            pf = q.order_by("-mtime_iso").first()
            if pf:
                pth = pf.path if os.path.isabs(pf.path) else os.path.join(file_root, (pf.rel_path or "").replace("/", os.sep))
                if os.path.isfile(pth):
                    return pth
        except Exception:
            pass
        return None

    table_data = [["Thumbnail", "Parent PN", "Rev", "Description", "Total Qty"]]
    for r in rows:
        qty = r.get("qty") or 0
        try:
            qty_s = f"{float(qty):g}"
        except Exception:
            qty_s = str(qty)
        pn = str(r.get("parent_pn") or "")
        rev = str(r.get("parent_rev") or "")
        desc = str(r.get("description") or "")
        img_cell = ""
        img_path = preview_png_path(pn, rev)
        if img_path:
            try:
                img_cell = RLImage(img_path, width=16*mm, height=16*mm)
                img_cell.hAlign = "CENTER"
            except Exception:
                img_cell = ""
        table_data.append([
            img_cell,
            pn,
            rev,
            desc,
            qty_s,
        ])

    col_widths = [20*mm, 32*mm, 16*mm, 84*mm, 18*mm]
    row_heights = [8*mm] + [18*mm] * (len(table_data) - 1)
    table = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
    ])
    table.setStyle(style)
    flowables.append(table)
    doc.build(flowables)
    return buf.getvalue()


def _cover_page_pdf(
    root_pn: str,
    root_rev: Optional[str],
    build_ts: Optional[datetime] = None,
    desc_override: Optional[str] = None,
    thumb_override: Optional[str] = None,
) -> Optional[bytes]:
    """Generate a minimal cover page with PN/REV, description, logo, and footer fields."""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfbase.pdfmetrics import stringWidth
        from reportlab.lib.utils import ImageReader
    except Exception:
        return None

    pdoc = _part_by(root_pn, root_rev)
    root_rev_clean = _clean_rev(root_rev) if root_rev is not None else _clean_rev(getattr(pdoc, "revision", "") if pdoc else "")
    attrs = harvest_part_attrs(pdoc) if pdoc else {}
    desc = _part_description(pdoc, attrs) if pdoc else (desc_override or "")

    proc_text = ", ".join([p for p in _part_processes(pdoc) if p])

    author = ""
    for key in ("author", "drawnby", "checkedby", "approved_by"):
        v = (attrs or {}).get(key)
        if v and str(v).strip():
            author = str(v).strip()
            break

    resolved = resolve_part_field_values(
        pdoc,
        ["material", "finish", "mass"],
        attrs=attrs,
        config=get_field_config(),
        extra={"part_number": root_pn, "revision": root_rev_clean},
    )
    material = str(resolved.get("material") or "").strip()
    finish = str(resolved.get("finish") or "").strip()
    mass = str(resolved.get("mass") or "").strip()

    related_file = ""
    try:
        q = PartFile.objects(part_number__iexact=root_pn, ext__iexact="pdf")
        if root_rev is not None or root_rev_clean:
            q = q.filter(revision__iexact=root_rev_clean)
        pf = q.order_by("rel_path").first()
        if pf:
            related_file = os.path.basename(pf.rel_path or pf.path or "")
    except Exception:
        related_file = ""

    ts = format_display_ts(build_ts)

    W, H = A4
    margin = 20 * mm
    logo_w = 30 * mm
    logo_h = 12 * mm
    header_w = W - (2 * margin) - logo_w - 6 * mm

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)

    # Logo top-right
    _draw_svg_or_png(c, W - margin - logo_w, H - margin - logo_h, logo_w, logo_h, "tinylogo.svg", "tinylogo.png")

    def _wrap_lines(text: str, font: str, size: float, max_w: float) -> List[str]:
        words = (text or "").split()
        if not words:
            return []
        lines: List[str] = []
        cur = ""
        for w in words:
            trial = (cur + " " + w).strip()
            if stringWidth(trial, font, size) <= max_w:
                cur = trial
            else:
                if cur:
                    lines.append(cur)
                cur = w
        if cur:
            lines.append(cur)
        return lines

    def _preview_png_path(pn: str, rev: str) -> Optional[str]:
        try:
            root = (current_app.config.get("FILE_ROOT_LOCAL") or current_app.config.get("FILES_LOCAL_ROOT") or "").rstrip("/\\")
            q = PartFile.objects(part_number__iexact=pn, ext_group="png", is_dwg=False)
            q = q.filter(revision__iexact=_norm_rev(rev))
            pf = q.order_by("-mtime_iso").first()
            if pf:
                pth = pf.path if os.path.isabs(pf.path) else os.path.join(root, (pf.rel_path or "").replace("/", os.sep))
                if os.path.isfile(pth):
                    return pth
        except Exception:
            pass
        return None

    # Header: PN/REV and description
    y = H - margin
    title = root_pn or ""
    if root_rev_clean:
        title = f"{root_pn}  REV {root_rev_clean}"
    c.setFont("Helvetica-Bold", 18)
    for line in _wrap_lines(title, "Helvetica-Bold", 18, max_w=header_w):
        y -= 7 * mm
        c.drawString(margin, y, line)

    if desc:
        c.setFont("Helvetica", 12)
        for line in _wrap_lines(desc, "Helvetica", 12, max_w=header_w):
            y -= 6 * mm
            c.drawString(margin, y, line)
    header_bottom = y

    # Footer: author, process, generated date, related file
    footer_items: List[str] = []
    if author:
        footer_items.append(f"Author: {author}")
    if proc_text:
        footer_items.append(f"Process: {proc_text}")
    if ts:
        footer_items.append(f"Generated: {ts}")
    if related_file:
        footer_items.append(f"Related file: {related_file}")
    if material:
        footer_items.append(f"Material: {material}")
    if finish:
        footer_items.append(f"Finish: {finish}")
    if mass:
        footer_items.append(f"Mass: {mass}")

    footer_lines: List[str] = []
    for item in footer_items:
        footer_lines.extend(_wrap_lines(item, "Helvetica", 10, max_w=W - 2 * margin))

    line_h = 5 * mm
    footer_top = margin + (len(footer_lines) * line_h)

    # Center image between header and footer
    try:
        img_path = thumb_override if (thumb_override and os.path.isfile(thumb_override)) else _preview_png_path(root_pn, root_rev_clean)
        if img_path:
            img_reader = ImageReader(img_path)
            iw, ih = img_reader.getSize()
            img_max_w = W - 2 * margin
            img_top = header_bottom - 8 * mm
            img_bottom = footer_top + 8 * mm
            img_max_h = max(0, img_top - img_bottom)
            if img_max_h > 10 * mm and iw > 0 and ih > 0:
                scale = min(img_max_w / iw, img_max_h / ih)
                draw_w = iw * scale
                draw_h = ih * scale
                x = (W - draw_w) / 2.0
                y = img_bottom + (img_max_h - draw_h) / 2.0
                c.drawImage(img_reader, x, y, width=draw_w, height=draw_h, preserveAspectRatio=True, mask="auto")
    except Exception:
        pass

    if footer_lines:
        y = margin + (len(footer_lines) - 1) * line_h
        c.setFont("Helvetica", 10)
        for line in footer_lines:
            c.drawString(margin, y, line)
            y -= line_h

    c.save()
    return buf.getvalue()


def _overlay_numbers_and_stamps(
    pdf_bytes: bytes,
    stamps: List[str],
    *,
    skip_first_page: bool = False,
    draw_page_numbers: bool = True,
) -> bytes:
    if not draw_page_numbers and not stamps:
        return pdf_bytes
    try:
        from reportlab.pdfgen import canvas
        from PyPDF2 import PdfReader, PdfWriter
    except Exception as exc:
        try:
            current_app.logger.error("PDF overlay failed: missing reportlab or PyPDF2")
        except Exception:
            pass
        raise RuntimeError("PDF numbering/stamping requires reportlab and PyPDF2.") from exc
    reader = PdfReader(io.BytesIO(pdf_bytes))
    writer = PdfWriter()
    # stamp images (optional)
    stamp_files = {
        'quote': 'quote_stamp.png',
        'classified': 'classified_stamp.png',
        'approved': 'approved_stamp.png',
        'wip': 'wip_stamp.png',
        'inprogress': None,
    }
    static_dir = os.path.join(os.path.dirname(__file__), '..', 'static', 'images')
    sel_paths = []
    for key, fn in stamp_files.items():
        if key in stamps:
            if fn is None:
                sel_paths.append((key, None))
                continue
            p = os.path.abspath(os.path.join(static_dir, fn))
            if not os.path.isfile(p):
                raise RuntimeError(f"Stamp asset missing: {fn}")
            sel_paths.append((key, p))

    total = len(reader.pages)
    for i, page in enumerate(reader.pages, start=1):
        mediabox = page.mediabox
        W = float(mediabox.width)
        H = float(mediabox.height)
        # build overlay for this page
        obuf = io.BytesIO()
        c = canvas.Canvas(obuf, pagesize=(W, H))
        # tiny footer bottom-center (logo/marker) on every page
        try:
            # 16mm x 6mm centered
            from reportlab.lib.units import mm
            _draw_svg_or_png(c, W/2 - 8*mm, 10, 16*mm, 6*mm, 'tinyfooter.svg', 'tinyfooter.png')
        except Exception:
            pass
        # page number bottom-right (skip on cover page if requested)
        if draw_page_numbers and not (skip_first_page and i == 1):
            c.setFont("Helvetica-Bold", 9)
            c.setFillGray(0.2)
            c.drawRightString(W - 20, 18, f"{i} / {total}")
        # stamps (semi-transparent)
        try:
            for key, spath in sel_paths:
                if key in ("approved","wip","inprogress"):
                    # center big watermark
                    c.saveState()
                    c.translate(W/2, H/2)
                    c.rotate(0)
                    c.setFillAlpha(0.2)
                    if key == "inprogress":
                        try:
                            size = max(24.0, min(W, H) * 0.12)
                            line_spacing = size * 1.15

                            c.setFont("Helvetica-Bold", size)
                            c.setFillColorRGB(0.8, 0.0, 0.0)

                            # Coordinates are centred around the current transformed origin.
                            c.drawCentredString(0, line_spacing / 2, "IN PROGRESS")
                            c.drawCentredString(0, -line_spacing / 2, "NOT APPROVED")
                        except Exception:
                            pass
                    else:
                        c.drawImage(spath, -W*0.25, -H*0.25, width=W*0.5, height=H*0.5, preserveAspectRatio=True, mask='auto')
                    c.restoreState()
                elif key in ("quote","classified"):
                    from reportlab.lib.units import mm
                    c.drawImage(spath, 15*mm, 40*mm, width=160, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass
        c.save()
        obuf.seek(0)
        # merge overlay
        from PyPDF2 import PdfReader as _PdfReader
        overlay = _PdfReader(obuf).pages[0]
        try:
            page.merge_page(overlay)
        except Exception:
            try:
                page.mergePage(overlay)  # older API
            except Exception:
                pass
        writer.add_page(page)
    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


def _merge_pdfs(paths: List[str]) -> Tuple[bytes, List[int]]:
    """Merge PDFs in order. Return (bytes, start_pages) where start_pages[i] is 1-based start page of doc i."""
    try:
        from PyPDF2 import PdfMerger, PdfReader
    except Exception:
        return b"", []
    merger = PdfMerger()
    starts: List[int] = []
    total = 1
    for p in paths:
        try:
            rdr = PdfReader(p)
            starts.append(total)
            total += len(rdr.pages)
            merger.append(rdr)
        except Exception:
            continue
    buf = io.BytesIO()
    merger.write(buf)
    merger.close()
    return (buf.getvalue(), starts)


def _merge_pdfs_filtered(paths: List[str], *, fp_tokens: List[str]) -> Tuple[bytes, List[int], List[int]]:
    """Merge PDFs in order, skipping pages whose label matches fp_tokens.
    Return (bytes, start_pages, kept_indices) where start_pages aligns to kept_indices.
    """
    try:
        from PyPDF2 import PdfReader, PdfWriter
    except Exception:
        return b"", [], []
    writer = PdfWriter()
    starts: List[int] = []
    kept: List[int] = []
    total = 1
    for idx, p in enumerate(paths):
        try:
            rdr = PdfReader(p)
        except Exception:
            continue
        labels = _page_labels(rdr) if fp_tokens else []
        start = total
        kept_pages = 0
        for page_idx, page in enumerate(rdr.pages):
            label = labels[page_idx] if labels and page_idx < len(labels) else ""
            if fp_tokens and label and _is_flat_pattern_page_label(label, fp_tokens):
                continue
            writer.add_page(page)
            kept_pages += 1
            total += 1
        if kept_pages:
            starts.append(start)
            kept.append(idx)
    buf = io.BytesIO()
    writer.write(buf)
    return (buf.getvalue(), starts, kept)


def _index_label_suffix(path: str, ext_group: str, seq: int) -> str:
    if (ext_group or "").lower() == "markup":
        return "markups"
    if (ext_group or "").lower() == "datasheet":
        return "datasheet"
    base = os.path.splitext(os.path.basename(path or ""))[0].lower()
    if "datasheet" in base:
        return "datasheet"
    if "drawing" in base or re.search(r"(?:^|[^a-z0-9])(?:dwg|drw)(?:[^a-z0-9]|$)", base):
        return "drawing"
    if "revision" in base or re.search(r"(?:^|[^a-z0-9])rev(?:[^a-z0-9]|$)", base):
        return "rev"
    return f"pdf {seq}"


def _index_label_for_item(
    item: Dict[str, object],
    desc_cache: Dict[Tuple[str, str], str],
    multi_by_pn: Dict[str, int],
    seq_by_pn: Dict[str, int],
) -> str:
    pn = str(item.get("pn") or "")
    rev = str(item.get("rev") or "")
    key = (pn, _norm_rev(rev))
    if key not in desc_cache:
        pdoc = _part_by(pn, _norm_rev(rev))
        desc_cache[key] = _part_description(pdoc) if pdoc else ""
    desc = desc_cache.get(key) or ""
    label = pn
    if desc:
        label = f"{pn} - {desc}"
    if multi_by_pn.get(pn, 0) > 1:
        seq_by_pn[pn] = seq_by_pn.get(pn, 0) + 1
        suffix = _index_label_suffix(str(item.get("path") or ""), str(item.get("ext_group") or ""), seq_by_pn[pn])
        if suffix:
            label = f"{label} ({suffix})" if label else suffix
    return label


def _index_pdf_standalone(
    pdf_items: List[Dict[str, object]],
    root_pn: Optional[str],
    root_rev: Optional[str],
    root_desc: Optional[str],
    *,
    build_ts: Optional[datetime] = None,
) -> Optional[bytes]:
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfbase.pdfmetrics import stringWidth
    except Exception:
        return None

    desc_cache: Dict[Tuple[str, str], str] = {}
    multi_by_pn: Dict[str, int] = {}
    for item in pdf_items:
        pn = str(item.get("pn") or "")
        multi_by_pn[pn] = multi_by_pn.get(pn, 0) + 1

    seq_by_pn: Dict[str, int] = {}
    labels = [
        _index_label_for_item(item, desc_cache, multi_by_pn, seq_by_pn)
        for item in pdf_items
    ]

    def _clip_line(text: str, font: str, size: float, max_w: float) -> str:
        text = text or ""
        if not text:
            return ""
        if stringWidth(text, font, size) <= max_w:
            return text
        suffix = "..."
        max_w = max(0.0, max_w - stringWidth(suffix, font, size))
        out = text
        while out and stringWidth(out, font, size) > max_w:
            out = out[:-1]
        return out.rstrip() + (suffix if out else "")

    def _wrap_lines(text: str, font: str, size: float, max_w: float, max_lines: int) -> List[str]:
        text = (text or "").strip()
        if not text or max_lines <= 0:
            return []
        words = text.split()
        lines: List[str] = []
        cur = ""
        truncated = False
        for w in words:
            cand = (cur + " " + w).strip()
            if stringWidth(cand, font, size) <= max_w:
                cur = cand
            else:
                if cur:
                    lines.append(cur)
                cur = w
                if len(lines) >= max_lines:
                    truncated = True
                    cur = ""
                    break
        if cur and len(lines) < max_lines:
            lines.append(cur)
        if truncated and lines:
            lines[-1] = _clip_line(lines[-1], font, size, max_w)
        return lines

    def _wrap_lines(text: str, font: str, size: float, max_w: float, max_lines: int) -> List[str]:
        text = (text or "").strip()
        if not text or max_lines <= 0:
            return []
        words = text.split()
        lines: List[str] = []
        cur = ""
        truncated = False
        for w in words:
            cand = (cur + " " + w).strip()
            if stringWidth(cand, font, size) <= max_w:
                cur = cand
            else:
                if cur:
                    lines.append(cur)
                cur = w
                if len(lines) >= max_lines:
                    truncated = True
                    cur = ""
                    break
        if cur and len(lines) < max_lines:
            lines.append(cur)
        if truncated and lines:
            lines[-1] = _clip_line(lines[-1], font, size, max_w)
        return lines

    def _wrap_lines(text: str, font: str, size: float, max_w: float, max_lines: int) -> List[str]:
        text = (text or "").strip()
        if not text or max_lines <= 0:
            return []
        words = text.split()
        lines: List[str] = []
        cur = ""
        truncated = False
        for w in words:
            cand = (cur + " " + w).strip()
            if stringWidth(cand, font, size) <= max_w:
                cur = cand
            else:
                if cur:
                    lines.append(cur)
                cur = w
                if len(lines) >= max_lines:
                    truncated = True
                    cur = ""
                    break
        if cur and len(lines) < max_lines:
            lines.append(cur)
        if truncated and lines:
            lines[-1] = _clip_line(lines[-1], font, size, max_w)
        return lines

    def _wrap_lines(text: str, font: str, size: float, max_w: float, max_lines: int) -> List[str]:
        text = (text or "").strip()
        if not text or max_lines <= 0:
            return []
        words = text.split()
        lines: List[str] = []
        cur = ""
        truncated = False
        for w in words:
            cand = (cur + " " + w).strip()
            if stringWidth(cand, font, size) <= max_w:
                cur = cand
            else:
                if cur:
                    lines.append(cur)
                cur = w
                if len(lines) >= max_lines:
                    truncated = True
                    cur = ""
                    break
        if cur and len(lines) < max_lines:
            lines.append(cur)
        if truncated and lines:
            lines[-1] = _clip_line(lines[-1], font, size, max_w)
        return lines

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    W, H = A4
    left_x = 20 * mm
    right_x = W - 20 * mm
    y = H - 20 * mm

    c.setFont("Helvetica-Bold", 18)
    c.drawString(left_x, y, "Index")
    y -= 8 * mm

    c.setFont("Helvetica", 10)
    stamp = format_display_ts(build_ts)
    if stamp:
        c.drawString(left_x, y, f"Generated: {stamp}")
        y -= 6 * mm
    if root_pn:
        root_rev_clean = _clean_rev(root_rev) if root_rev is not None else ""
        c.drawString(left_x, y, f"Part: {root_pn}    Rev: {root_rev_clean}")
        y -= 6 * mm
    if root_desc:
        avail_w = right_x - left_x
        words = str(root_desc).split()
        cur = ""
        while words and y > (H / 2):
            t = (cur + " " + words[0]).strip()
            if stringWidth(t, "Helvetica", 10) <= avail_w:
                cur = t
                words.pop(0)
            else:
                if cur:
                    c.drawString(left_x, y, f"Description: {cur}")
                    y -= 6 * mm
                    cur = ""
                else:
                    c.drawString(left_x, y, f"Description: {t[:60]}")
                    y -= 6 * mm
                    words.pop(0)
                    break
        if cur and y > (H / 2):
            c.drawString(left_x, y, f"Description: {cur}")
            y -= 6 * mm
    y -= 2 * mm

    if not labels:
        c.drawString(left_x, y, "No PDFs found.")
        c.save()
        return buf.getvalue()

    c.setFont("Helvetica", 10)
    max_w = right_x - left_x
    for label in labels:
        if y < 20 * mm:
            c.showPage()
            y = H - 20 * mm
            c.setFont("Helvetica", 10)
        c.drawString(left_x, y, _clip_line(label, "Helvetica", 10, max_w))
        y -= 6 * mm

    c.save()
    return buf.getvalue()


def build_docpack(opts: DocPackOptions) -> Tuple[str, bytes, str]:
    """Return (filename, bytes, mime). Currently returns a ZIP by default, unless only a single PDF binder is requested.
    """
    build_ts = get_display_dt(getattr(opts, "build_ts", None))
    base_stub = (opts.output_name or "").strip()
    if not base_stub:
        base_stub = (opts.root_pn or "docpack").strip()
        if opts.root_rev:
            rev_stub = _clean_rev(opts.root_rev)
            if rev_stub:
                base_stub = f"{base_stub}_{rev_stub}"
    else:
        base_stub = os.path.splitext(base_stub)[0]

    def _flat_or_override(*args, **kwargs):
        if opts.flat_override is not None:
            return list(opts.flat_override)
        return _flatten_bom(*args, **kwargs)

    # 1) Build BOM
    # define terminal processes for "consumed" logic
    consumed_terminals = ["welding", "purchase", "machine"]
    flat = _flat_or_override(
        opts.root_pn,
        opts.root_rev,
        full=(opts.depth != "top"),
        include_consumed=bool(opts.include_consumed),
        terminal_processes=consumed_terminals,
    )

    # 2) Filter parts by classification/process
    filtered_flat: List[Tuple[str,str,float]] = []
    # fabrication_pack convenience filters
    fab_procs = {
        "welding",
        "lasercut",
        "profile cut",
        "folding",
        "rolling",
        "cutting",
        "machine",
        "3d laser",
        "casting",
    }
    fab_cut_procs = {
        "lasercut",
        "profile cut",
        "folding",
        "rolling",
        "cutting",
        "plasma",
        "waterjet",
    }
    fab_all: Optional[List[Tuple[str, str, float]]] = None
    fab_scope: Optional[List[Tuple[str, str, float]]] = None
    force_proc_filter = False
    if getattr(opts, "fabrication_pack", False):
        # restrict to fab processes
        opts.process_mode = "selected"
        opts.processes = list(sorted(fab_procs))
        force_proc_filter = True
        # scope of supply: exclude consumed parts
        opts.include_consumed = False
        # include docs for fabrication pack (exclude PNGs per request)
        opts.file_types = ["dxf", "step", "pdf"]
        opts.want_selected_files = True
        opts.want_excel_bom = True
        opts.want_pdf_binder = True
        fab_full = (opts.depth != "top")
        # all fabrication parts (include consumed)
        fab_all = _flatten_bom(
            opts.root_pn,
            opts.root_rev,
            full=fab_full,
            include_consumed=True,
            terminal_processes=consumed_terminals,
        )
        # scope of supply: stop at fabrication parts
        fab_scope = _flatten_bom(
            opts.root_pn,
            opts.root_rev,
            full=fab_full,
            include_consumed=False,
            terminal_processes=list(fab_procs),
        )

    def _filter_flat(source: Iterable[Tuple[str, str, float]]) -> List[Tuple[str, str, float]]:
        out: List[Tuple[str, str, float]] = []
        for pn, rev, qty in source:
            p = _part_by(pn, rev)
            if not p:
                continue
            if not _passes_classified_filter(p, opts.classified_filter):
                continue
            if not _passes_process_filter(p, opts.processes, opts.process_mode):
                continue
            out.append((pn, rev, qty))
        return out

    flat_source = fab_all if fab_all is not None else flat
    filtered_flat = _filter_flat(flat_source)

    # 3) Collect files
    # include datasheets in binder if requested (affects file_types for binder only)
    root_part = _part_by(opts.root_pn, None) if opts.root_rev is None else None
    root_rev_resolved = _clean_rev(opts.root_rev) if opts.root_rev is not None else _clean_rev(getattr(root_part, "revision", "") if root_part else "")
    root_desc = ""
    try:
        pdoc = root_part if root_part is not None else _part_by(opts.root_pn, root_rev_resolved)
        root_desc = _part_description(pdoc)
    except Exception:
        pass
    if not root_desc and opts.root_desc_override:
        root_desc = opts.root_desc_override
    chosen_files = _collect_files(filtered_flat + [(opts.root_pn, root_rev_resolved, 1.0)], opts.file_types)
    markup_pairs = [(opts.root_pn, root_rev_resolved)] + [(pn, rev) for pn, rev, _qty in filtered_flat]
    need_markup_documents = bool(opts.want_markup_files or opts.want_markup_report or (opts.want_pdf_binder and opts.binder_add_markups))
    markup_documents = markup_documents_for_pairs(markup_pairs) if need_markup_documents else []
    output_count = 0
    if opts.want_selected_files:
        output_count += 1
    if opts.want_excel_bom:
        output_count += 1
    if opts.want_index_pdf:
        output_count += 1
    if opts.want_visual_list:
        output_count += 1
    if opts.want_cover_page:
        output_count += 1
    if opts.want_whereused_report:
        output_count += 1
    if getattr(opts, "want_hardware_summary", False):
        output_count += 1
    if opts.want_markup_report:
        output_count += 1
    if opts.want_markup_files:
        output_count += 1
    if getattr(opts, "fabrication_pack", False):
        output_count += 2
    if opts.want_pdf_binder:
        output_count += 1
    want_zip = bool(opts.want_selected_files or opts.want_markup_files) or output_count != 1
    if not want_zip and opts.want_excel_bom:
        want_zip = True
    # Precompute PDF body items for binder/index
    pairs: List[Tuple[str, str]] = []
    pdf_items: List[Dict[str, object]] = []
    pdf_paths: List[str] = []
    missing_pdf_rows: List[Dict[str, object]] = []
    if opts.want_pdf_binder or opts.want_index_pdf:
        uniq_children: Dict[Tuple[str, str], None] = {}
        for pn, rev, _ in filtered_flat:
            key = (pn, _norm_rev(rev))
            uniq_children[key] = None
        root_key = (opts.root_pn, root_rev_resolved)
        if root_key in uniq_children:
            uniq_children.pop(root_key, None)
        ordered_children = sorted(list(uniq_children.keys()), key=lambda t: (t[0] or "", t[1] or ""))
        pairs = [root_key] + ordered_children
        try:
            pdf_items, pdf_paths = _pdf_items_for_pairs(
                pairs,
                include_datasheets=bool(getattr(opts, "binder_add_datasheets", False)),
                include_flat_patterns=bool(getattr(opts, "binder_include_flat_patterns", False)),
            )
        except Exception:
            current_app.logger.exception(
                "docpack: failed to collect body PDFs for %s:%s; binder body will be empty",
                opts.root_pn, opts.root_rev or "",
            )
        if opts.want_pdf_binder:
            try:
                missing_pdf_rows = _missing_pdf_rows(pairs, current_app.config.get("FILE_ROOT_LOCAL") or "")
            except Exception:
                current_app.logger.exception(
                    "docpack: failed to compute missing-file diagnostics for %s:%s",
                    opts.root_pn, opts.root_rev or "",
                )
    markup_tmpdir: Optional[str] = None
    if opts.want_pdf_binder and opts.binder_add_markups and markup_documents:
        markup_tmpdir = tempfile.mkdtemp(prefix="tinymrp-markups-")
        markup_by_pair = {(doc.part_number.lower(), _norm_rev(doc.revision)): doc for doc in markup_documents}
        reordered_items: List[Dict[str, object]] = []
        reordered_paths: List[str] = []
        for pair_pn, pair_rev in pairs:
            pair_key = (str(pair_pn or "").lower(), _norm_rev(pair_rev))
            for item, path in zip(pdf_items, pdf_paths):
                if (str(item.get("pn") or "").lower(), _norm_rev(str(item.get("rev") or ""))) == pair_key:
                    reordered_items.append(item)
                    reordered_paths.append(path)
            markup_doc = markup_by_pair.get(pair_key)
            if markup_doc:
                markup_path = os.path.join(markup_tmpdir, markup_doc.filename)
                with open(markup_path, "wb") as handle:
                    handle.write(markup_doc.pdf_bytes)
                reordered_items.append({
                    "pn": markup_doc.part_number,
                    "rev": markup_doc.revision,
                    "path": markup_path,
                    "ext_group": "markup",
                })
                reordered_paths.append(markup_path)
        pdf_items, pdf_paths = reordered_items, reordered_paths
    # 4) Build payloads (ZIP container)
    zip_buf = io.BytesIO()
    z: Optional[zipfile.ZipFile] = None
    if want_zip:
        z = zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED)

    # Excel BOM
    if opts.want_excel_bom:
        occ_map = {} if opts.flat_override is not None else _bom_occurrences(opts.root_pn, opts.root_rev, include_consumed=bool(opts.include_consumed), terminal_processes=["welding","purchase","machine"])
        root_key = (opts.root_pn, _norm_rev(root_rev_resolved))
        if root_key not in occ_map and opts.flat_override is None:
            occ_map[root_key] = [("+", 1.0)]
        full_flat = _flat_or_override(
            opts.root_pn,
            opts.root_rev,
            full=True,
            include_consumed=bool(opts.include_consumed),
            terminal_processes=["welding","purchase","machine"],
        )
        full_qty_map = { (pn, _norm_rev(rev)): qty for pn, rev, qty in full_flat }
        excel_flat = list(filtered_flat)
        root_key_norm = (opts.root_pn or "").strip().lower(), _norm_rev(root_rev_resolved)
        if opts.flat_override is None and not any(((pn or "").strip().lower(), _norm_rev(rev)) == root_key_norm for pn, rev, _ in excel_flat):
            excel_flat.insert(0, (opts.root_pn, root_rev_resolved, 1.0))
        xlsx = _excel_bom_bytes(
            opts.root_pn,
            opts.root_rev,
            excel_flat,
            occ_map,
            full_qty_map,
            include_all_fields=bool(getattr(opts, "excel_all_fields", False)),
            field_ids=list(getattr(opts, "excel_field_ids", None) or []),
        )
        bom_name = build_output_name(f"{base_stub}_BOM", "xlsx", max_len=96, include_time=False, now=build_ts)
        z.writestr(bom_name, xlsx)

    # Selected files
    if opts.want_selected_files and chosen_files:
        root = (current_app.config.get("FILE_ROOT_LOCAL") or "").rstrip("/\\")
        for f in chosen_files:
            rp = f.rel_path.replace("\\","/") if f.rel_path else os.path.basename(f.path)
            abs_path = f.path if os.path.isabs(f.path) else os.path.join(root, rp.replace("/", os.sep))
            arcname = rp
            try:
                z.write(abs_path, arcname)
            except Exception:
                continue

    if opts.want_markup_files and markup_documents:
        for document in markup_documents:
            z.writestr(f"markups/{document.filename}", document.pdf_bytes)

    if want_zip:
        z.writestr(
            "_file_coverage.txt",
            _file_coverage_report(filtered_flat, current_app.config.get("FILE_ROOT_LOCAL") or ""),
        )

    # Precompute visual list rows when needed (standalone or binder sections)
    vis_filtered_all: List[Tuple[str, str, float]] = []
    vis_filtered_visual: List[Tuple[str, str, float]] = []
    need_vis = bool(opts.want_visual_list) or bool(getattr(opts, "want_hardware_summary", False)) or (
        bool(opts.want_pdf_binder) and (bool(opts.binder_add_visual_list) or bool(opts.binder_add_hardware_summary))
    )
    if need_vis:
        if fab_all is not None:
            vis_source = fab_all
        else:
            vis_source = _flat_or_override(
                opts.root_pn,
                opts.root_rev,
                full=True,
                include_consumed=bool(opts.include_consumed),
                terminal_processes=["welding", "purchase", "machine"],
            )
        vis_filtered_all = _filter_flat(vis_source)
        vis_filtered_visual = [
            (pn, rev, qty)
            for pn, rev, qty in vis_filtered_all
            if not _is_hardware_part(pn, rev)
        ]

    # Hardware Summary PDF (standalone or binder section)
    hardware_rows: List[Dict[str, object]] = []
    hardware_pdf: Optional[bytes] = None
    if bool(getattr(opts, "want_hardware_summary", False)) or bool(opts.binder_add_hardware_summary):
        try:
            hardware_rows = _hardware_summary_rows(vis_filtered_all + [(opts.root_pn, root_rev_resolved, 1.0)])
            hardware_pdf = _hardware_summary_pdf_with_empty(
                hardware_rows,
                allow_empty=bool(getattr(opts, "want_hardware_summary", False)),
                root_pn=opts.root_pn,
                root_rev=root_rev_resolved,
                root_desc=root_desc,
                build_ts=build_ts,
            )
            if hardware_rows and not hardware_pdf:
                raise RuntimeError("Failed to build Hardware Summary PDF. Ensure reportlab is installed.")
        except RuntimeError:
            raise
        except Exception:
            pass

    embed_hw_summary = bool(
        hardware_rows
        and (
            (opts.want_visual_list and opts.want_hardware_summary)
            or (opts.binder_add_visual_list and opts.binder_add_hardware_summary)
        )
    )
    visual_hw_rows = hardware_rows if embed_hw_summary else None
    visual_hw_pdf = hardware_pdf if embed_hw_summary else None

    if getattr(opts, "want_hardware_summary", False):
        if not hardware_pdf:
            raise RuntimeError("Failed to build Hardware Summary PDF. Ensure reportlab is installed.")
        hw_name = build_output_name(f"{base_stub}_HardwareSummary", "pdf", max_len=96, include_time=False, now=build_ts)
        if want_zip:
            z.writestr(hw_name, hardware_pdf)
        else:
            return (hw_name, hardware_pdf, "application/pdf")

    if opts.want_markup_report:
        if not markup_documents:
            raise RuntimeError("No drawing markups are available for the selected parts.")
        markup_report = combine_markup_documents(
            markup_documents,
            root_label=f"{opts.root_pn} {root_rev_resolved}".strip() + (f" - {root_desc}" if root_desc else ""),
            build_ts=build_ts,
        )
        markup_report_name = build_output_name(f"{base_stub}_MarkupReport", "pdf", max_len=96, include_time=False, now=build_ts)
        if want_zip:
            z.writestr(markup_report_name, markup_report)
        else:
            return (markup_report_name, markup_report, "application/pdf")

    # Index PDF (standalone)
    if opts.want_index_pdf:
        idx_pdf = _index_pdf_standalone(
            pdf_items,
            opts.root_pn,
            root_rev_resolved,
            root_desc,
            build_ts=build_ts,
        )
        if not idx_pdf:
            raise RuntimeError("Failed to build Index PDF. Ensure reportlab is installed.")
        idx_name = build_output_name(f"{base_stub}_Index", "pdf", max_len=96, include_time=False, now=build_ts)
        if want_zip:
            z.writestr(idx_name, idx_pdf)
        else:
            return (idx_name, idx_pdf, "application/pdf")

    # Visual list PDF (standalone)
    if opts.want_visual_list:
        # Standalone visual list includes root header/cell and omits hardware children
        vis_pdf = _visual_list_pdf(
            vis_filtered_visual,
            opts.root_pn,
            root_rev_resolved,
            build_ts=build_ts,
            hardware_rows=visual_hw_rows,
            hardware_pdf_bytes=visual_hw_pdf,
            root_desc=root_desc,
            root_thumb_override=opts.root_thumb_override,
        )
        if not vis_pdf:
            raise RuntimeError("Failed to build Visual List PDF. Ensure reportlab and qrcode are installed.")
        if want_zip:
            vis_name = build_output_name(f"{base_stub}_VisualList", "pdf", max_len=96, include_time=False, now=build_ts)
            z.writestr(vis_name, vis_pdf)
        else:
            # single artifact path (rare)
            vis_name = build_output_name(f"{base_stub}_VisualList", "pdf", max_len=96, include_time=False, now=build_ts)
            return (vis_name, vis_pdf, "application/pdf")

    cover_pdf = None
    if opts.want_cover_page or (opts.want_pdf_binder and opts.binder_add_cover):
        cover_pdf = _cover_page_pdf(
            opts.root_pn,
            opts.root_rev,
            build_ts=build_ts,
            desc_override=opts.root_desc_override,
            thumb_override=opts.root_thumb_override,
        )
        if not cover_pdf:
            raise RuntimeError("Failed to build binder cover page. Ensure reportlab is installed.")
    if opts.want_cover_page and cover_pdf:
        cover_name = build_output_name(f"{base_stub}_Cover", "pdf", max_len=96, include_time=False, now=build_ts)
        if want_zip:
            z.writestr(cover_name, cover_pdf)
        else:
            return (cover_name, cover_pdf, "application/pdf")

    whereused_pdf = None
    if opts.want_whereused_report or (opts.want_pdf_binder and opts.binder_add_whereused):
        where_rows = _whereused_rows(opts.root_pn, opts.root_rev)
        whereused_pdf = _whereused_report_pdf(where_rows, opts.root_pn, opts.root_rev, build_ts=build_ts)
        if not whereused_pdf:
            raise RuntimeError("Failed to build Where-Used report. Ensure reportlab is installed.")
    if opts.want_whereused_report and whereused_pdf:
        where_name = build_output_name(f"{base_stub}_WhereUsed", "pdf", max_len=96, include_time=False, now=build_ts)
        if want_zip:
            z.writestr(where_name, whereused_pdf)
        else:
            return (where_name, whereused_pdf, "application/pdf")

    # Fabrication pack extras: scope-of-supply + cut/fold summary
    if getattr(opts, "fabrication_pack", False):
        scope_source = _filter_flat(fab_scope or [])
        scope_rows = _scope_supply_rows(scope_source)
        scope_pdf = _scope_supply_pdf(scope_rows, "Scope of Supply")
        if scope_rows and not scope_pdf:
            raise RuntimeError("Failed to build Scope of Supply PDF. Ensure reportlab is installed.")
        if scope_pdf:
            scope_name = build_output_name(f"{base_stub}_ScopeOfSupply", "pdf", max_len=96, include_time=False, now=build_ts)
            z.writestr(scope_name, scope_pdf)

        cut_rows_src: List[Tuple[str, str, float]] = []
        for pn, rev, qty in (fab_all or []):
            p = _part_by(pn, rev)
            if not p:
                continue
            if not _passes_classified_filter(p, opts.classified_filter):
                continue
            if not _passes_process_filter(p, opts.processes, opts.process_mode):
                continue
            if set(_part_processes(p)) & fab_cut_procs:
                cut_rows_src.append((pn, rev, qty))
        cut_rows = _cut_fold_summary_rows(cut_rows_src)
        cut_pdf = _cut_fold_summary_pdf(cut_rows, "Laser / Fold Summary")
        if cut_rows and not cut_pdf:
            raise RuntimeError("Failed to build Laser/Fold Summary PDF. Ensure reportlab is installed.")
        if cut_pdf:
            cut_name = build_output_name(f"{base_stub}_LaserFoldSummary", "pdf", max_len=96, include_time=False, now=build_ts)
            z.writestr(cut_name, cut_pdf)

    # PDF binder (with index & page numbers)
    if opts.want_pdf_binder:
        # Gather PDFs independent of UI file filter (always include PDFs if present)
        # pdf_items/pdf_paths already prepared above to keep binder + index consistent.

        # Preface: cover, index, and optional sections
        preface_bytes: List[Tuple[str, bytes]] = []
        cover = cover_pdf if opts.binder_add_cover else None
        if opts.binder_add_cover:
            if not cover:
                raise RuntimeError("Failed to build binder cover page. Ensure reportlab is installed.")
            preface_bytes.append(("Cover.pdf", cover))
        vis_pdf = None
        if opts.binder_add_visual_list:
            vis_pdf = _visual_list_pdf(
                vis_filtered_visual,
                opts.root_pn,
                root_rev_resolved,
                build_ts=build_ts,
                hardware_rows=visual_hw_rows,
                hardware_pdf_bytes=visual_hw_pdf,
                root_desc=root_desc,
                root_thumb_override=opts.root_thumb_override,
            )
            if not vis_pdf:
                raise RuntimeError("Failed to build Visual List PDF. Ensure reportlab and qrcode are installed.")
            preface_bytes.append(("VisualList.pdf", vis_pdf))
        if opts.binder_add_whereused:
            if whereused_pdf is None:
                where_rows = _whereused_rows(opts.root_pn, opts.root_rev)
                whereused_pdf = _whereused_report_pdf(where_rows, opts.root_pn, opts.root_rev, build_ts=build_ts)
                if not whereused_pdf:
                    raise RuntimeError("Failed to build Where-Used report. Ensure reportlab is installed.")
            preface_bytes.append(("WhereUsed.pdf", whereused_pdf))

        if opts.binder_add_hardware_summary and not embed_hw_summary:
            if hardware_pdf and hardware_rows:
                preface_bytes.append(("HardwareSummary.pdf", hardware_pdf))

        if missing_pdf_rows:
            missing_files_pdf = _missing_files_pdf(missing_pdf_rows)
            if missing_files_pdf:
                preface_bytes.append(("MissingFiles.pdf", missing_files_pdf))

        # Merge body first to measure page counts (optionally filter flat pattern pages)
        binder_pdf_items = list(pdf_items)
        binder_pdf_paths = list(pdf_paths)
        body_bytes = b""
        body_starts: List[int] = []
        if binder_pdf_paths:
            if not opts.binder_include_flat_patterns:
                fp_tokens = _flat_pattern_page_tokens()
                if fp_tokens:
                    body_bytes, body_starts, kept_idx = _merge_pdfs_filtered(
                        binder_pdf_paths,
                        fp_tokens=fp_tokens,
                    )
                    binder_pdf_items = [binder_pdf_items[i] for i in kept_idx]
                    binder_pdf_paths = [binder_pdf_paths[i] for i in kept_idx]
                else:
                    body_bytes, body_starts = _merge_pdfs(binder_pdf_paths)
            else:
                body_bytes, body_starts = _merge_pdfs(binder_pdf_paths)

        # Build index in two passes to account for its own page count, include Visual Summary entry, dot leaders, and metadata
        # Skip index entirely if no harvested PDFs were found
        try:
            if not opts.binder_add_index:
                raise RuntimeError("skip_index")
            if not binder_pdf_paths:
                raise RuntimeError("no_body_pdfs")
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import mm
            from reportlab.pdfbase.pdfmetrics import stringWidth
            from PyPDF2 import PdfReader

            desc_cache: Dict[Tuple[str, str], str] = {}
            multi_by_pn: Dict[str, int] = {}
            for item in binder_pdf_items:
                pn = str(item.get("pn") or "")
                multi_by_pn[pn] = multi_by_pn.get(pn, 0) + 1

            # preface counts excluding index
            cover_pages = 0
            vis_pages = 0
            where_pages = 0
            hardware_pages = 0
            missing_pages = 0
            for name, b in preface_bytes:
                lc = name.lower()
                if lc.startswith("cover"):
                    cover_pages += len(PdfReader(io.BytesIO(b)).pages)
                elif lc.startswith("visual"):
                    vis_pages += len(PdfReader(io.BytesIO(b)).pages)
                elif lc.startswith("where"):
                    where_pages += len(PdfReader(io.BytesIO(b)).pages)
                elif lc.startswith("hardware"):
                    hardware_pages += len(PdfReader(io.BytesIO(b)).pages)
                elif lc.startswith("missingfiles"):
                    missing_pages += len(PdfReader(io.BytesIO(b)).pages)

            # Helper to build the index PDF with an assumed index page count
            def _build_index(assumed_idx_pages: int) -> bytes:
                seq_by_pn: Dict[str, int] = {}
                idx_io = io.BytesIO()
                c = canvas.Canvas(idx_io, pagesize=A4)
                W, H = A4
                left_x = 20*mm
                right_x = W - 20*mm
                y = H - 20*mm
                # Title
                c.setFont("Helvetica-Bold", 18)
                c.drawString(left_x, y, "Index")
                y -= 8*mm
                # Binder metadata
                c.setFont("Helvetica", 10)
                now = format_display_ts(build_ts)
                pdoc = _part_by(opts.root_pn, root_rev_resolved)
                desc = _part_description(pdoc) if pdoc else ''
                c.drawString(left_x, y, f"Generated: {now}"); y -= 6*mm
                c.drawString(left_x, y, f"Part: {opts.root_pn}    Rev: {root_rev_resolved}"); y -= 6*mm
                if desc:
                    # wrap description
                    avail_w = right_x - left_x
                    words = desc.split(); cur='';
                    while words and y > (H/2):
                        t = (cur + ' ' + words[0]).strip()
                        if stringWidth(t, "Helvetica", 10) <= avail_w:
                            cur = t; words.pop(0)
                        else:
                            if cur:
                                c.drawString(left_x, y, f"Description: {cur}")
                                y -= 6*mm; cur=''
                            else:
                                # extremely long word, truncate
                                c.drawString(left_x, y, f"Description: {t[:60]}…"); y-=6*mm; words.pop(0)
                                break
                    if cur and y > (H/2):
                        c.drawString(left_x, y, f"Description: {cur}"); y -= 6*mm
                y -= 2*mm
                # Entries
                c.setFont("Helvetica", 10)
                dot_w = stringWidth('.', 'Helvetica', 10)
                def _entry(name: str, page_no: int):
                    nonlocal y
                    if y < 20*mm:
                        c.showPage(); y = H - 20*mm; c.setFont("Helvetica", 10)
                    label = name
                    left_w = stringWidth(label, 'Helvetica', 10)
                    page_s = str(page_no)
                    right_w = stringWidth(page_s, 'Helvetica', 10)
                    dots_area = max(0, right_x - (left_x + left_w) - right_w - 6)
                    n_dots = int(dots_area / max(dot_w, 0.1))
                    c.drawString(left_x, y, label)
                    if n_dots > 0:
                        c.drawString(left_x + left_w + 3, y, '.' * n_dots)
                    c.drawRightString(right_x, y, page_s)
                    y -= 6*mm

                # Visual Summary entry first (always before body)
                if vis_pdf:
                    vis_start = cover_pages + assumed_idx_pages + 1
                    _entry("Visual Summary", vis_start)
                if where_pages:
                    where_start = cover_pages + assumed_idx_pages + (vis_pages if vis_pdf else 0) + 1
                    _entry("Where Used", where_start)
                if hardware_pages:
                    hw_start = cover_pages + assumed_idx_pages + (vis_pages if vis_pdf else 0) + where_pages + 1
                    _entry("Hardware Summary", hw_start)
                if missing_pages:
                    missing_start = cover_pages + assumed_idx_pages + (vis_pages if vis_pdf else 0) + where_pages + hardware_pages + 1
                    _entry("Missing Source Documents", missing_start)
                # Body entries: PN + Description
                pre_body_offset = cover_pages + assumed_idx_pages + (vis_pages if vis_pdf else 0) + where_pages + hardware_pages + missing_pages
                for item, start in zip(binder_pdf_items, body_starts):
                    label = _index_label_for_item(item, desc_cache, multi_by_pn, seq_by_pn)
                    _entry(label, pre_body_offset + start)
                c.save()
                return idx_io.getvalue()

            # First pass (assume 1 page), then measure and rebuild
            idx_b1 = _build_index(assumed_idx_pages=1)
            idx_pages = len(PdfReader(io.BytesIO(idx_b1)).pages)
            idx_b2 = _build_index(assumed_idx_pages=idx_pages)

            # Place index after cover
            insert_pos = 1 if cover else 0
            preface_bytes.insert(insert_pos, ("Index.pdf", idx_b2))
        except Exception:
            pass

        # If we have body PDFs, compute part -> first page map and rebuild VisualList with page numbers
        if binder_pdf_paths and opts.binder_add_visual_list:
            try:
                # Compute cover/index/visual page counts now that index is finalized
                from PyPDF2 import PdfReader
                cover_pages = 0
                vis_pages = 0
                index_pages = 0
                where_pages = 0
                hardware_pages = 0
                missing_pages = 0
                for name, b in preface_bytes:
                    lc = name.lower()
                    if lc.startswith('cover'):
                        cover_pages += len(PdfReader(io.BytesIO(b)).pages)
                    elif lc.startswith('index'):
                        index_pages += len(PdfReader(io.BytesIO(b)).pages)
                    elif lc.startswith('visual'):
                        vis_pages += len(PdfReader(io.BytesIO(b)).pages)
                    elif lc.startswith('where'):
                        where_pages += len(PdfReader(io.BytesIO(b)).pages)
                    elif lc.startswith('hardware'):
                        hardware_pages += len(PdfReader(io.BytesIO(b)).pages)
                    elif lc.startswith('missingfiles'):
                        missing_pages += len(PdfReader(io.BytesIO(b)).pages)
                pre_body_offset = cover_pages + index_pages + vis_pages + where_pages + hardware_pages + missing_pages
                # Map absolute path -> start page
                start_by_path = {p: s for p, s in zip(binder_pdf_paths, body_starts)}
                # For each part, pick earliest starting page among its PDFs
                first_page: Dict[Tuple[str,str], int] = {}
                for item in binder_pdf_items:
                    ap = str(item.get("path") or "")
                    sp = start_by_path.get(ap)
                    if sp is None:
                        continue
                    key = (str(item.get("pn") or ""), str(item.get("rev") or ""))
                    val = pre_body_offset + int(sp)
                    if key not in first_page or val < first_page[key]:
                        first_page[key] = val
                # Rebuild visual list with page numbers (ensure full-BOM aggregated quantities)
                vis_full2 = _flat_or_override(
                    opts.root_pn,
                    opts.root_rev,
                    full=True,
                    include_consumed=bool(opts.include_consumed),
                    terminal_processes=["welding","purchase","machine"],
                )
                vis_filtered2: List[Tuple[str,str,float]] = []
                for pn, rev, qty in vis_full2:
                    p = _part_by(pn, rev)
                    if not p:
                        continue
                    if not _passes_classified_filter(p, opts.classified_filter):
                        continue
                    if not _passes_process_filter(p, opts.processes, opts.process_mode):
                        continue
                    if _is_hardware_part(pn, rev):
                        continue
                    vis_filtered2.append((pn, rev, qty))
                vis_pdf2 = _visual_list_pdf(
                    vis_filtered2,
                    opts.root_pn,
                    root_rev_resolved,
                    page_map=first_page,
                    build_ts=build_ts,
                    hardware_rows=visual_hw_rows,
                    hardware_pdf_bytes=visual_hw_pdf,
                    root_desc=root_desc,
                    root_thumb_override=opts.root_thumb_override,
                )
                # Replace existing VisualList in preface_bytes
                for i, (nm, _) in enumerate(preface_bytes):
                    if nm.lower().startswith('visual'):
                        preface_bytes[i] = (nm, vis_pdf2)
                        break
            except Exception:
                pass

        # Merge final binder: preface (cover + index + visual list) + body
        all_segments: List[bytes] = [b for _, b in preface_bytes]
        if body_bytes:
            all_segments.append(body_bytes)
        final_pdf: bytes = b""
        if all_segments:
            seg_paths = []
            tmpdir = tempfile.mkdtemp()
            try:
                for i, b in enumerate(all_segments):
                    p = os.path.join(tmpdir, f"seg_{i}.pdf")
                    with open(p, 'wb') as fh:
                        fh.write(b)
                    seg_paths.append(p)
                merged_pdf, _ = _merge_pdfs(seg_paths)
                final_pdf = merged_pdf or b""
            finally:
                pass

        # Always apply page numbers (skip cover page); include stamps if requested
        stamps = []
        if opts.stamp_quote:
            stamps.append('quote')
        if opts.stamp_confidential:
            stamps.append('classified')
        status_stamp = None
        if opts.stamp_approved:
            status_stamp = 'approved'
        elif opts.stamp_wip:
            status_stamp = 'wip'
        elif opts.stamp_inprogress:
            status_stamp = 'inprogress'
        if status_stamp:
            stamps.append(status_stamp)
        if final_pdf:
            final_pdf = _overlay_numbers_and_stamps(
                final_pdf,
                stamps,
                skip_first_page=bool(cover),
                draw_page_numbers=bool(opts.binder_page_numbers),
            )

        if final_pdf:
            if want_zip:
                binder_name = build_output_name(f"{base_stub}_Binder", "pdf", max_len=96, include_time=False, now=build_ts)
                z.writestr(binder_name, final_pdf)
            else:
                binder_name = build_output_name(f"{base_stub}_Binder", "pdf", max_len=96, include_time=False, now=build_ts)
                return (binder_name, final_pdf, "application/pdf")
    # Finish ZIP
    if z is not None:
        z.close()
    data = zip_buf.getvalue()
    name = build_output_name(f"{base_stub}_DocPack", "zip", max_len=96, include_time=False, now=build_ts)
    return (name, data, "application/zip")

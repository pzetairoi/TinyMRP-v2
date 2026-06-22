import io
import zipfile

from openpyxl import load_workbook

from app.models.auth import Role, User
from app.models.bom import BOMLink
from app.models.part import Part
from app.services.docpacks import DocPackOptions, build_docpack
from app.services.field_config import field_requires_runtime_scan, query_paths_for_field, save_field_config


def _login(client, user):
    with client.session_transaction() as sess:
        sess["_user_id"] = user.get_id()


def _admin_user():
    role = Role(name="admin").save()
    return User(
        email="admin@example.com",
        password="test",
        active=True,
        fs_uniquifier="admin-field-config",
        roles=[role],
    ).save()


def _config_payload():
    return {
        "builtin_fields": [
            {"id": "description", "label": "Description", "source_path": "attrs.summary_text"},
            {"id": "material", "label": "Material", "source_path": "attrs.mat_code"},
        ],
        "custom_fields": [
            {"id": "legacy_code", "label": "Legacy Code", "source_path": "attrs.legacy_code", "data_type": "text"},
        ],
        "contexts": {
            "parts_list": {
                "allowed_field_ids": [
                    "thumbnail",
                    "part_number",
                    "revision",
                    "description",
                    "material",
                    "legacy_code",
                ],
                "default_field_ids": [
                    "part_number",
                    "description",
                    "material",
                    "legacy_code",
                ],
            },
            "part_detail_summary": {
                "allowed_field_ids": ["description", "material", "legacy_code"],
                "default_field_ids": ["description", "material", "legacy_code"],
            },
            "excel_bom": {
                "allowed_field_ids": [
                    "part_number",
                    "revision",
                    "description",
                    "total_qty",
                    "legacy_code",
                ],
                "default_field_ids": [
                    "part_number",
                    "revision",
                    "description",
                    "total_qty",
                    "legacy_code",
                ],
            },
        },
    }


def test_field_config_api_and_user_preferences(client):
    admin = _admin_user()
    _login(client, admin)

    resp = client.put("/api/admin/field-config", json=_config_payload())
    assert resp.status_code == 200

    get_resp = client.get("/api/field-config")
    assert get_resp.status_code == 200
    data = get_resp.get_json()
    assert data["ok"] is True
    assert data["permissions"]["can_admin"] is True
    parts_ctx = data["config"]["contexts"]["parts_list"]
    assert "legacy_code" in parts_ctx["allowed_field_ids"]
    desc_field = next(field for field in data["config"]["fields"] if field["id"] == "description")
    assert desc_field["source_path"] == "attrs.summary_text"

    prefs_resp = client.put(
        "/api/me/settings",
        json={"field_preferences": {"contexts": {"parts_list": {"field_ids": ["legacy_code", "part_number"]}}}},
    )
    assert prefs_resp.status_code == 200
    prefs_data = prefs_resp.get_json()
    assert prefs_data["settings"]["field_preferences"]["contexts"]["parts_list"]["field_ids"] == ["legacy_code", "part_number"]


def test_parts_and_part_detail_use_custom_field_mapping(client):
    admin = _admin_user()
    _login(client, admin)
    client.put("/api/admin/field-config", json=_config_payload())

    part = Part(
        part_number="CFG-100",
        revision="A",
        description="Legacy Description",
        attrs={"summary_text": "Mapped Description", "mat_code": "SS304", "legacy_code": "LEG-100"},
    ).save()

    list_resp = client.post("/api/parts_lazy", json={"first": 0, "rows": 25, "filters": {}})
    assert list_resp.status_code == 200
    rows = list_resp.get_json()["data"]
    row = next(item for item in rows if item["part_number"] == "CFG-100")
    assert row["description"] == "Mapped Description"
    assert row["material"] == "SS304"
    assert row["legacy_code"] == "LEG-100"

    detail_resp = client.get(f"/api/part_detail?pn={part.part_number}&rev={part.revision}")
    assert detail_resp.status_code == 200
    detail = detail_resp.get_json()
    assert detail["part"]["description"] == "Mapped Description"
    assert detail["part"]["field_values"]["legacy_code"] == "LEG-100"
    assert detail["part"]["field_values"]["material"] == "SS304"


def test_excel_bom_respects_selected_field_ids(app):
    root = Part(part_number="ASM-FLD", revision="", description="Root").save()
    child = Part(
        part_number="CMP-FLD",
        revision="B",
        description="Legacy Child",
        attrs={"summary_text": "Mapped Child", "legacy_code": "LC-77"},
    ).save()
    BOMLink(parent_pn=root.part_number, parent_rev="", child_pn=child.part_number, child_rev="B", qty=3).save()

    with app.app_context():
        save_field_config(_config_payload())
        _, data, mime = build_docpack(
            DocPackOptions(
                root_pn=root.part_number,
                root_rev="",
                depth="full",
                include_consumed=True,
                want_excel_bom=True,
                excel_field_ids=["part_number", "revision", "description", "legacy_code", "total_qty"],
                want_selected_files=False,
                want_pdf_binder=False,
                want_visual_list=False,
            )
        )

    assert mime == "application/zip"
    zf = zipfile.ZipFile(io.BytesIO(data))
    bom_name = next(name for name in zf.namelist() if name.lower().endswith(".xlsx"))
    wb = load_workbook(io.BytesIO(zf.read(bom_name)))
    ws = wb.active
    header = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    data_row = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=3, max_row=3))]

    assert header == ["Part Number", "Revision", "Description", "Legacy Code", "Total Qty"]
    assert "Mapped Child" in data_row
    assert "LC-77" in data_row


def test_custom_fields_normalize_source_paths_and_raw_attr_keys_across_views(client, app):
    admin = _admin_user()
    _login(client, admin)

    resp = client.put(
        "/api/admin/field-config",
        json={
            "builtin_fields": [
                {"id": "description", "label": "Description", "source_path": "attrs.Summary Text"},
            ],
            "custom_fields": [
                {"id": "Legacy Code", "label": "Legacy Code", "source_path": "", "data_type": "text"},
                {"id": "ETA-Date", "label": "ETA Date", "source_path": "attrs.ETA-Date", "data_type": "text"},
            ],
            "contexts": {
                "parts_list": {
                    "allowed_field_ids": ["part_number", "description", "legacy_code", "eta_date"],
                    "default_field_ids": ["part_number", "description", "legacy_code", "eta_date"],
                },
                "part_detail_summary": {
                    "allowed_field_ids": ["description", "legacy_code", "eta_date"],
                    "default_field_ids": ["description", "legacy_code", "eta_date"],
                },
                "excel_bom": {
                    "allowed_field_ids": ["part_number", "revision", "description", "total_qty", "legacy_code", "eta_date"],
                    "default_field_ids": ["part_number", "revision", "description", "total_qty", "legacy_code", "eta_date"],
                },
            },
        },
    )
    assert resp.status_code == 200

    config_resp = client.get("/api/field-config")
    assert config_resp.status_code == 200
    config_data = config_resp.get_json()["config"]
    fields_by_id = {field["id"]: field for field in config_data["fields"]}
    assert fields_by_id["description"]["source_path"] == "attrs.summary_text"
    assert fields_by_id["legacy_code"]["source_path"] == "attrs.legacy_code"
    assert fields_by_id["eta_date"]["source_path"] == "attrs.eta_date"

    root = Part(part_number="ASM-SET", revision="", description="Root").save()
    child = Part(
        part_number="CMP-SET",
        revision="A",
        description="Child",
        attrs={"Summary Text": "Mapped Child", "Legacy Code": "LEG-100", "ETA-Date": "2026-07-01"},
    ).save()
    BOMLink(parent_pn=root.part_number, parent_rev="", child_pn=child.part_number, child_rev=child.revision, qty=2).save()

    list_resp = client.post("/api/parts_lazy", json={"first": 0, "rows": 25, "filters": {}})
    assert list_resp.status_code == 200
    rows = list_resp.get_json()["data"]
    row = next(item for item in rows if item["part_number"] == child.part_number)
    assert row["description"] == "Mapped Child"
    assert row["legacy_code"] == "LEG-100"
    assert row["eta_date"] == "2026-07-01"

    filter_resp = client.post(
        "/api/parts_lazy",
        json={
            "first": 0,
            "rows": 25,
            "filters": {
                "description": {"value": "Mapped Child"},
                "legacy_code": {"value": "LEG-100"},
                "eta_date": {"value": "2026-07-01"},
            },
        },
    )
    assert filter_resp.status_code == 200
    filtered_rows = filter_resp.get_json()["data"]
    assert [item["part_number"] for item in filtered_rows] == [child.part_number]

    detail_resp = client.get(f"/api/part_detail?pn={child.part_number}&rev={child.revision}")
    assert detail_resp.status_code == 200
    detail = detail_resp.get_json()
    assert detail["part"]["description"] == "Mapped Child"
    assert detail["part"]["field_values"]["legacy_code"] == "LEG-100"
    assert detail["part"]["field_values"]["eta_date"] == "2026-07-01"

    with app.app_context():
        _, data, mime = build_docpack(
            DocPackOptions(
                root_pn=root.part_number,
                root_rev="",
                depth="full",
                include_consumed=True,
                want_excel_bom=True,
                excel_field_ids=["part_number", "revision", "description", "legacy_code", "eta_date", "total_qty"],
                want_selected_files=False,
                want_pdf_binder=False,
                want_visual_list=False,
            )
        )

    assert mime == "application/zip"
    zf = zipfile.ZipFile(io.BytesIO(data))
    bom_name = next(name for name in zf.namelist() if name.lower().endswith(".xlsx"))
    wb = load_workbook(io.BytesIO(zf.read(bom_name)))
    ws = wb.active
    header = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    data_row = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=3, max_row=3))]

    assert header == ["Part Number", "Revision", "Description", "Legacy Code", "ETA Date", "Total Qty"]
    assert "LEG-100" in data_row
    assert "2026-07-01" in data_row


def test_field_candidate_endpoint_returns_normalized_attr_catalog(client):
    admin = _admin_user()
    _login(client, admin)

    Part(
        part_number="CAT-100",
        revision="A",
        attrs={
            "ETA-Date": "2026-07-01",
            "Location": "Rack 1",
            "Inspection Required": True,
            "Supplier Portal": "https://example.com/specs/cat-100",
            "process2": "paint",
        },
    ).save()
    Part(
        part_number="CAT-200",
        revision="A",
        attrs={
            "eta date": "2026-07-02",
            "location": "Rack 2",
            "Inspection Required": "false",
        },
    ).save()

    resp = client.get("/api/admin/field-config/candidates")
    assert resp.status_code == 200

    candidates = resp.get_json()["candidates"]
    by_id = {item["id"]: item for item in candidates}

    assert by_id["eta_date"]["label"] == "ETA Date"
    assert by_id["eta_date"]["source_path"] == "attrs.eta_date"
    assert by_id["eta_date"]["part_count"] == 2
    assert set(by_id["eta_date"]["raw_keys"]) == {"ETA-Date", "eta date"}

    assert by_id["location"]["source_path"] == "attrs.location"
    assert by_id["location"]["part_count"] == 2
    assert by_id["inspection_required"]["data_type"] == "boolean"
    assert by_id["supplier_portal"]["data_type"] == "link"

    assert "description" not in by_id
    assert "material" not in by_id
    assert "process2" not in by_id
    assert "processes" not in by_id


def test_filters_use_resolved_values_for_custom_and_remapped_fields(client):
    admin = _admin_user()
    _login(client, admin)

    client.put(
        "/api/admin/field-config",
        json={
            "builtin_fields": [
                {"id": "description", "label": "Description", "source_path": "attrs.Summary Text"},
            ],
            "custom_fields": [
                {"id": "legacy_code", "label": "Legacy Code", "source_path": "attrs.Legacy Code", "data_type": "text"},
            ],
            "contexts": {
                "parts_list": {
                    "allowed_field_ids": ["part_number", "description", "legacy_code"],
                    "default_field_ids": ["part_number", "description", "legacy_code"],
                }
            },
        },
    )

    Part(part_number="FLT-RAW-1", revision="A", description="Legacy", attrs={"Summary Text": "Alpha Bravo", "Legacy Code": "LC-100"}).save()
    Part(part_number="FLT-RAW-2", revision="A", description="Legacy", attrs={"Summary Text": "Charlie Delta", "Legacy Code": "LC-200"}).save()

    desc_resp = client.post(
        "/api/parts_lazy",
        json={"first": 0, "rows": 25, "sortField": "description", "filters": {"description": {"value": "Alpha"}}},
    )
    assert desc_resp.status_code == 200
    assert [row["part_number"] for row in desc_resp.get_json()["data"]] == ["FLT-RAW-1"]

    custom_resp = client.post(
        "/api/parts_lazy",
        json={"first": 0, "rows": 25, "sortField": "legacy_code", "filters": {"legacy_code": {"value": "LC-200"}}},
    )
    assert custom_resp.status_code == 200
    assert [row["part_number"] for row in custom_resp.get_json()["data"]] == ["FLT-RAW-2"]

    config = client.get("/api/field-config").get_json()["config"]
    assert query_paths_for_field("description", config) == ["field_values__description"]
    assert query_paths_for_field("legacy_code", config) == ["field_values__legacy_code"]
    assert field_requires_runtime_scan("description", config) is False
    assert field_requires_runtime_scan("legacy_code", config) is False

    global_resp = client.post(
        "/api/parts_lazy",
        json={"first": 0, "rows": 25, "filters": {"global": {"value": "LC-200"}}},
    )
    assert global_resp.status_code == 200
    assert [row["part_number"] for row in global_resp.get_json()["data"]] == ["FLT-RAW-2"]


def test_parts_lazy_filters_custom_text_without_materialized_field_values(client):
    admin = _admin_user()
    _login(client, admin)

    client.put(
        "/api/admin/field-config",
        json={
            "custom_fields": [
                {"id": "legacy_code", "label": "Legacy Code", "source_path": "attrs.Legacy Code", "data_type": "text"},
            ],
            "contexts": {
                "parts_list": {
                    "allowed_field_ids": ["part_number", "description", "legacy_code"],
                    "default_field_ids": ["part_number", "description", "legacy_code"],
                }
            },
        },
    )

    Part(
        part_number="STALE-100",
        revision="A",
        description="Legacy",
        attrs={"Legacy Code": "LEG-100"},
    ).save(sync_materialized=False)
    Part(
        part_number="STALE-200",
        revision="A",
        description="Legacy",
        attrs={"Legacy Code": "LEG-200"},
    ).save(sync_materialized=False)

    stale = Part.objects(part_number="STALE-100", revision="A").first()
    assert stale is not None
    assert dict(stale.field_values or {}) == {}

    resp = client.post(
        "/api/parts_lazy",
        json={"first": 0, "rows": 25, "filters": {"legacy_code": {"value": "LEG-100"}}},
    )
    assert resp.status_code == 200
    rows = resp.get_json()["data"]
    assert [row["part_number"] for row in rows] == ["STALE-100"]
    assert rows[0]["legacy_code"] == "LEG-100"


def test_admin_can_rebuild_search_fields_for_existing_parts(client):
    admin = _admin_user()
    _login(client, admin)

    legacy = Part(
        part_number="REB-100",
        revision="A",
        description="Legacy part",
        attrs={"Legacy Code": "LEG-100", "datasheet": "vendor-file.pdf"},
    ).save()

    save_resp = client.put(
        "/api/admin/field-config",
        json={
            "custom_fields": [
                {"id": "legacy_code", "label": "Legacy Code", "source_path": "attrs.Legacy Code", "data_type": "text"},
            ],
            "contexts": {
                "parts_list": {
                    "allowed_field_ids": ["part_number", "description", "legacy_code", "has_datasheet"],
                    "default_field_ids": ["part_number", "description", "legacy_code", "has_datasheet"],
                }
            },
        },
    )
    assert save_resp.status_code == 200

    rebuild_resp = client.post("/api/admin/field-config/rebuild-search-fields")
    assert rebuild_resp.status_code == 200
    report = rebuild_resp.get_json()["report"]
    assert report["updated"] >= 1
    assert report["errors"] == 0

    legacy.reload()
    assert legacy.field_values["legacy_code"] == "LEG-100"
    assert legacy.has_datasheet is True
    assert "datasheet" in (legacy.file_groups or [])

    list_resp = client.post(
        "/api/parts_lazy",
        json={
            "first": 0,
            "rows": 25,
            "filters": {
                "legacy_code": {"value": "LEG-100"},
                "has_datasheet": {"value": True},
            },
        },
    )
    assert list_resp.status_code == 200
    assert [row["part_number"] for row in list_resp.get_json()["data"]] == [legacy.part_number]


def test_admin_can_rebuild_canonical_fields_using_custom_aliases(client):
    admin = _admin_user()
    _login(client, admin)

    part = Part(
        part_number="CAN-100",
        revision="A",
        description="Alias Part",
        attrs={"comments": "LASERCUT", "secondprocess": "MACHINE"},
    ).save()

    save_resp = client.put(
        "/api/admin/field-config",
        json={
            "canonical_aliases": [
                {
                    "field_id": "process",
                    "aliases": ["process", "processes", "comments", "secondprocess", "thirdprocess"],
                }
            ]
        },
    )
    assert save_resp.status_code == 200

    rebuild_resp = client.post("/api/admin/field-config/rebuild-canonical-fields")
    assert rebuild_resp.status_code == 200
    report = rebuild_resp.get_json()["report"]
    assert report["updated"] >= 1

    part.reload()
    assert part.attrs["comments"] == "LASERCUT"
    assert part.attrs["secondprocess"] == "MACHINE"
    assert (part.processes or []) == ["lasercut", "machine"]
    assert (part.canonical or {}).get("processes") == ["lasercut", "machine"]

    list_resp = client.post("/api/parts_lazy", json={"first": 0, "rows": 25, "filters": {"process": {"value": "machine"}}})
    assert list_resp.status_code == 200
    rows = list_resp.get_json()["data"]
    assert [row["part_number"] for row in rows] == [part.part_number]

    detail_resp = client.get(f"/api/part_detail?pn={part.part_number}&rev={part.revision}")
    assert detail_resp.status_code == 200
    detail = detail_resp.get_json()
    assert detail["part"]["processes"] == ["lasercut", "machine"]
    assert detail["part"]["process"] == "lasercut, machine"

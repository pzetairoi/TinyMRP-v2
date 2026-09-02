"""The depth choice must mean the same thing in every output of a pack.

"This part + its children" (depth="top") covers the root assembly and the
components directly under it, and stops there. It used to be honoured by the
selected files, the Excel BOM and the binder body, while the visual summary and
the hardware summary walked the whole tree - so a first-level pack still listed
grandchildren. These tests hold every output to the one scope.
"""

import io
import os
import zipfile

from openpyxl import load_workbook
from pypdf import PdfReader
from reportlab.pdfgen import canvas

from app.models.artifact import PartFile
from app.models.bom import BOMLink
from app.models.part import Part
from app.services.docpacks import DocPackOptions, build_docpack


ROOT = "DEPTH-ROOT-001"
CHILD = "DEPTH-CHILD-002"
GRAND = "DEPTH-GRAND-003"
GREAT = "DEPTH-GREAT-004"
CHILD_HW = "DEPTH-HW-CHILD"
GRAND_HW = "DEPTH-HW-GRAND"


def _write_pdf(path: str, label: str) -> None:
    c = canvas.Canvas(path)
    c.drawString(100, 750, label)
    c.showPage()
    c.save()


def _build_tree(file_root: str) -> None:
    """ROOT -> CHILD -> GRAND -> GREAT, with hardware hung at two levels."""
    Part(part_number=ROOT, revision="1", description="ROOT ASSEMBLY",
         processes=["welding"]).save()
    Part(part_number=CHILD, revision="1", description="CHILD SUB ASSEMBLY",
         processes=["welding"]).save()
    Part(part_number=GRAND, revision="1", description="GRANDCHILD PLATE",
         processes=["lasercut"]).save()
    Part(part_number=GREAT, revision="1", description="GREAT GRANDCHILD PLATE",
         processes=["lasercut"]).save()
    Part(part_number=CHILD_HW, revision="", description="FIRST LEVEL SCREW",
         processes=["hardware"]).save()
    Part(part_number=GRAND_HW, revision="", description="DEEP LEVEL SCREW",
         processes=["hardware"]).save()

    BOMLink(parent_pn=ROOT, parent_rev="1", child_pn=CHILD, child_rev="1", qty=2).save()
    BOMLink(parent_pn=ROOT, parent_rev="1", child_pn=CHILD_HW, child_rev="", qty=4).save()
    BOMLink(parent_pn=CHILD, parent_rev="1", child_pn=GRAND, child_rev="1", qty=3).save()
    BOMLink(parent_pn=CHILD, parent_rev="1", child_pn=GRAND_HW, child_rev="", qty=8).save()
    BOMLink(parent_pn=GRAND, parent_rev="1", child_pn=GREAT, child_rev="1", qty=5).save()

    for pn in (ROOT, CHILD, GRAND, GREAT):
        pdf_path = os.path.join(file_root, f"{pn}.pdf")
        _write_pdf(pdf_path, f"{pn} DRAWING")
        PartFile(part_number=pn, revision="1", ext_group="pdf", ext="pdf",
                 rel_path=f"{pn}.pdf", path=pdf_path).save()


def _opts(depth: str, **kwargs) -> DocPackOptions:
    base = dict(
        root_pn=ROOT,
        root_rev="1",
        depth=depth,
        include_consumed=True,
        file_types=["pdf"],
        want_selected_files=False,
    )
    base.update(kwargs)
    return DocPackOptions(**base)


def _zip_of(data: bytes) -> zipfile.ZipFile:
    return zipfile.ZipFile(io.BytesIO(data))


def _pdf_text(data: bytes) -> str:
    reader = PdfReader(io.BytesIO(data))
    return "\n".join((page.extract_text() or "") for page in reader.pages)


def test_first_level_visual_summary_stops_at_immediate_children(app, tmp_path):
    app.config["FILE_ROOT_LOCAL"] = str(tmp_path)
    _build_tree(str(tmp_path))

    with app.app_context():
        _name, data, mime = build_docpack(_opts("top", want_visual_list=True))

    assert mime == "application/pdf"
    text = _pdf_text(data)
    assert ROOT in text, "the root is the subject of the pack and must appear"
    assert CHILD in text, "an immediate child belongs in a first-level pack"
    assert GRAND not in text, "a grandchild must not appear in a first-level pack"
    assert GREAT not in text, "a great-grandchild must not appear in a first-level pack"


def test_full_depth_visual_summary_still_reaches_every_level(app, tmp_path):
    app.config["FILE_ROOT_LOCAL"] = str(tmp_path)
    _build_tree(str(tmp_path))

    with app.app_context():
        _name, data, _mime = build_docpack(_opts("full", want_visual_list=True))

    text = _pdf_text(data)
    for pn in (CHILD, GRAND, GREAT):
        assert pn in text, f"{pn} missing from a full-BOM visual summary"


def test_first_level_hardware_summary_stops_at_immediate_children(app, tmp_path):
    app.config["FILE_ROOT_LOCAL"] = str(tmp_path)
    _build_tree(str(tmp_path))

    with app.app_context():
        _name, data, _mime = build_docpack(_opts("top", want_hardware_summary=True))

    text = _pdf_text(data)
    assert CHILD_HW in text, "hardware directly under the root belongs in the pack"
    assert GRAND_HW not in text, "hardware inside a sub-assembly is below the chosen depth"


def test_first_level_binder_sections_and_body_share_one_scope(app, tmp_path):
    app.config["FILE_ROOT_LOCAL"] = str(tmp_path)
    _build_tree(str(tmp_path))

    with app.app_context():
        _name, data, mime = build_docpack(
            _opts(
                "top",
                want_pdf_binder=True,
                binder_add_cover=True,
                binder_add_index=True,
                binder_add_visual_list=True,
                binder_add_hardware_summary=True,
                binder_page_numbers=True,
            )
        )

    assert mime == "application/pdf"
    text = _pdf_text(data)
    assert f"{ROOT} DRAWING" in text
    assert f"{CHILD} DRAWING" in text
    assert f"{GRAND} DRAWING" not in text, "grandchild drawing merged into a first-level binder"
    assert f"{GREAT} DRAWING" not in text
    # The visual list and index sections are rendered into this same binder, so
    # a leak there shows up as the part number appearing without its drawing.
    assert GRAND not in text, "a binder section still lists grandchildren"
    assert GREAT not in text


def test_first_level_excel_bom_rows_and_levels_stay_at_one_level(app, tmp_path):
    app.config["FILE_ROOT_LOCAL"] = str(tmp_path)
    _build_tree(str(tmp_path))

    with app.app_context():
        _name, data, _mime = build_docpack(_opts("top", want_excel_bom=True))

    with _zip_of(data) as z:
        bom_name = next(n for n in z.namelist() if n.lower().endswith(".xlsx"))
        wb = load_workbook(io.BytesIO(z.read(bom_name)))
    ws = wb.active
    header = [str(c.value or "").strip().lower() for c in ws[1]]
    pn_col = header.index("part number")
    rows = [
        [c.value for c in row]
        for row in ws.iter_rows(min_row=2)
        if str(row[pn_col].value or "").strip()
    ]
    listed = {str(r[pn_col]).strip() for r in rows}

    assert ROOT in listed and CHILD in listed and CHILD_HW in listed
    assert GRAND not in listed, "grandchild row in a first-level Excel BOM"
    assert GREAT not in listed

    if "level" in header:
        level_col = header.index("level")
        for row in rows:
            level = str(row[level_col] or "")
            assert level.count(".") <= 1, (
                f"level path {level!r} describes a row deeper than the chosen depth"
            )


def test_first_level_selected_files_exclude_deeper_drawings(app, tmp_path):
    app.config["FILE_ROOT_LOCAL"] = str(tmp_path)
    _build_tree(str(tmp_path))

    with app.app_context():
        _name, data, _mime = build_docpack(_opts("top", want_selected_files=True))

    with _zip_of(data) as z:
        names = "\n".join(z.namelist())
    assert f"{ROOT}.pdf" in names and f"{CHILD}.pdf" in names
    assert f"{GRAND}.pdf" not in names
    assert f"{GREAT}.pdf" not in names


def test_first_level_excel_columns_ignore_deeper_uses_of_the_same_part(app, tmp_path):
    """A part used at the first level AND inside a sub-assembly.

    The Level and Total Qty columns are computed from a separate BOM walk than
    the rows are. That walk ignored the depth choice, so a first-level pack
    reported this screw at level "+.02, +.01.02" with a quantity that counted
    the 8 hidden inside the sub-assembly.
    """
    app.config["FILE_ROOT_LOCAL"] = str(tmp_path)
    _build_tree(str(tmp_path))
    # The first-level screw is also consumed inside the sub-assembly.
    BOMLink(parent_pn=CHILD, parent_rev="1", child_pn=CHILD_HW, child_rev="", qty=8).save()

    with app.app_context():
        _name, data, _mime = build_docpack(_opts("top", want_excel_bom=True))

    with _zip_of(data) as z:
        bom_name = next(n for n in z.namelist() if n.lower().endswith(".xlsx"))
        wb = load_workbook(io.BytesIO(z.read(bom_name)))
    ws = wb.active
    header = [str(c.value or "").strip().lower() for c in ws[1]]
    pn_col, level_col, total_col = (
        header.index("part number"),
        header.index("level"),
        header.index("total qty"),
    )
    row = next(
        r for r in ws.iter_rows(min_row=2, values_only=True)
        if str(r[pn_col] or "").strip() == CHILD_HW
    )

    assert str(row[level_col]) == "+.02", (
        f"level {row[level_col]!r} names a position below the chosen depth"
    )
    assert float(row[total_col]) == 4.0, (
        "total qty counted the uses inside the sub-assembly the pack excludes"
    )

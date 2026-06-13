import io
import zipfile

from openpyxl import load_workbook

from app.models.artifact import PartFile
from app.models.part import Part
from app.models.bom import BOMLink
from app.services.docpacks import DocPackOptions, build_docpack
from app.services.field_config import save_field_config


def test_excel_bom_uses_total_qty_header(app):
    root = Part(part_number="ASM-200", revision="", description="Root").save()
    child = Part(part_number="C-1", revision="", description="Child").save()
    BOMLink(parent_pn=root.part_number, parent_rev="", child_pn=child.part_number, child_rev="", qty=2).save()

    opts = DocPackOptions(
        root_pn=root.part_number,
        root_rev="",
        depth="full",
        include_consumed=True,
        want_excel_bom=True,
        want_selected_files=False,
        want_pdf_binder=False,
        want_visual_list=False,
    )

    with app.app_context():
        name, data, mime = build_docpack(opts)

    assert mime == "application/zip"
    zf = zipfile.ZipFile(io.BytesIO(data))
    bom_name = next(n for n in zf.namelist() if n.lower().endswith(".xlsx"))
    wb = load_workbook(io.BytesIO(zf.read(bom_name)))
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    header = [str(c.value or "").strip().lower() for c in header_row]

    assert "total qty" in header
    assert all("approved" not in h for h in header)


def test_excel_bom_all_fields_toggle(app):
    root = Part(part_number="ASM-300", revision="", description="Root").save()
    child = Part(part_number="C-2", revision="", description="Child", attrs={"custom_field": "X"}).save()
    BOMLink(parent_pn=root.part_number, parent_rev="", child_pn=child.part_number, child_rev="", qty=1).save()

    opts = DocPackOptions(
        root_pn=root.part_number,
        root_rev="",
        depth="full",
        include_consumed=True,
        want_excel_bom=True,
        excel_all_fields=False,
        want_selected_files=False,
        want_pdf_binder=False,
        want_visual_list=False,
    )

    with app.app_context():
        _, data_main, _ = build_docpack(opts)

    zf = zipfile.ZipFile(io.BytesIO(data_main))
    bom_name = next(n for n in zf.namelist() if n.lower().endswith(".xlsx"))
    wb = load_workbook(io.BytesIO(zf.read(bom_name)))
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    header_main = [str(c.value or "").strip().lower() for c in header_row]

    assert "custom_field" not in header_main

    opts.excel_all_fields = True
    with app.app_context():
        _, data_full, _ = build_docpack(opts)

    zf_full = zipfile.ZipFile(io.BytesIO(data_full))
    bom_name_full = next(n for n in zf_full.namelist() if n.lower().endswith(".xlsx"))
    wb_full = load_workbook(io.BytesIO(zf_full.read(bom_name_full)))
    ws_full = wb_full.active
    header_row_full = next(ws_full.iter_rows(min_row=1, max_row=1))
    header_full = [str(c.value or "").strip().lower() for c in header_row_full]

    assert "custom_field" in header_full


def test_excel_bom_file_availability_fields_use_real_coverage(app):
    root = Part(part_number="ASM-400", revision="A", description="Root").save()
    child = Part(part_number="C-4", revision="B", description="Child").save()
    BOMLink(parent_pn=root.part_number, parent_rev=root.revision, child_pn=child.part_number, child_rev=child.revision, qty=1).save()
    PartFile(
        part_number=child.part_number,
        revision=child.revision,
        ext_group="stl",
        ext="stl",
        rel_path="stl/C-4_REV_B.stl",
        path="C:/vault/stl/C-4_REV_B.stl",
    ).save()

    with app.app_context():
        save_field_config(
            {
                "contexts": {
                    "excel_bom": {
                        "allowed_field_ids": ["part_number", "revision", "description", "total_qty", "has_stl", "has_pdf"],
                        "default_field_ids": ["part_number", "revision", "description", "has_stl", "has_pdf", "total_qty"],
                    }
                }
            }
        )
        _, data, mime = build_docpack(
            DocPackOptions(
                root_pn=root.part_number,
                root_rev=root.revision,
                depth="full",
                include_consumed=True,
                want_excel_bom=True,
                excel_field_ids=["part_number", "revision", "description", "has_stl", "has_pdf", "total_qty"],
                want_selected_files=False,
                want_pdf_binder=False,
                want_visual_list=False,
            )
        )

    assert mime == "application/zip"
    zf = zipfile.ZipFile(io.BytesIO(data))
    bom_name = next(n for n in zf.namelist() if n.lower().endswith(".xlsx"))
    wb = load_workbook(io.BytesIO(zf.read(bom_name)))
    ws = wb.active
    header = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    child_row = [c.value for c in next(ws.iter_rows(min_row=3, max_row=3))]

    stl_idx = header.index("Has STL")
    pdf_idx = header.index("Has PDF")
    assert child_row[stl_idx] is True
    assert child_row[pdf_idx] is False

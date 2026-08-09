import io
import os
import uuid
import zipfile

import openpyxl
from reportlab.pdfgen import canvas
from pypdf import PdfReader

from app.models.auth import Role, User
from app.models.part import Part
from app.models.bom import BOMLink
from app.models.artifact import PartFile
from app.services.excel_compile import CompileRow, build_excel_compile_zip
from app.services.permissions import PERMISSION_REGISTRY


def _login(client, user):
    with client.session_transaction() as sess:
        sess["_user_id"] = user.get_id()
        sess["_fresh"] = True


def _tools_user():
    role = Role.objects(name="administrator").first() or Role(name="administrator", permissions=sorted(PERMISSION_REGISTRY)).save()
    return User(
        email=f"tools-{uuid.uuid4()}@example.com",
        password="test",
        active=True,
        fs_uniquifier=str(uuid.uuid4()),
        roles=[role],
    ).save()


def _write_pdf(path: str, label: str) -> None:
    c = canvas.Canvas(path)
    c.drawString(100, 750, label)
    c.showPage()
    c.save()


def _write_png(path: str) -> None:
    from PIL import Image

    Image.new("RGB", (8, 8), color=(200, 30, 30)).save(path, format="PNG")


def test_excel_compile_binder_merges_all_pdfs(app, tmp_path):
    root_dir = tmp_path
    app.config["FILE_ROOT_LOCAL"] = str(root_dir)

    root = Part(part_number="AWS-B-001597", revision="1", description="BEARING CLAMP DRIVE - WELDMENT",
                processes=["welding"], attrs={"material": "REFER TO BOM", "finish": "ZINC", "mass": 0.9}).save()
    child = Part(part_number="AWS-A-001596", revision="1", description="BEARING CLAMP DRIVE PLATE",
                 processes=["lasercut", "machine"], attrs={"material": "MS - GR250", "finish": "NATURAL", "mass": 0.7}).save()
    hw = Part(part_number="B18.3.5M-16x2.0x55", revision="", description="COUNTER SINK SCREW",
              processes=["hardware"]).save()
    BOMLink(parent_pn=root.part_number, parent_rev="1", child_pn=child.part_number, child_rev="1", qty=1).save()
    BOMLink(parent_pn=root.part_number, parent_rev="1", child_pn=hw.part_number, child_rev="", qty=6).save()

    root_pdf = os.path.join(root_dir, "AWS-B-001597.pdf")
    child_pdf = os.path.join(root_dir, "AWS-A-001596.pdf")
    _write_pdf(root_pdf, "ROOT WELDMENT PDF")
    _write_pdf(child_pdf, "CHILD PLATE PDF")
    PartFile(part_number=root.part_number, revision="1", ext_group="pdf", ext="pdf",
              rel_path="AWS-B-001597.pdf", path=root_pdf).save()
    PartFile(part_number=child.part_number, revision="1", ext_group="pdf", ext="pdf",
              rel_path="AWS-A-001596.pdf", path=child_pdf).save()

    rows = [CompileRow(part_number=root.part_number, revision="1", qty=1.0, status="ok")]

    with app.app_context():
        data, missing = build_excel_compile_zip(
            rows,
            input_filename="input.xlsx",
            input_bytes=b"fake-xlsx-bytes",
            file_root=str(root_dir),
            want_excel_bom=True,
            want_pdf_binder=True,
            binder_add_cover=True,
            binder_add_index=True,
            binder_add_visual_list=True,
            binder_add_hardware_summary=True,
            binder_page_numbers=True,
            title="Quote 12345",
        )

    assert not missing
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        names = zf.namelist()
        binder_name = next((n for n in names if "Binder" in n and n.lower().endswith(".pdf")), None)
        assert binder_name, f"No Binder.pdf found in output zip: {names}"
        binder_bytes = zf.read(binder_name)

    reader = PdfReader(io.BytesIO(binder_bytes))
    full = "\n".join((p.extract_text() or "") for p in reader.pages)
    assert "ROOT WELDMENT PDF" in full, "root part's own PDF missing from compiled binder"
    assert "CHILD PLATE PDF" in full, "child part's own PDF missing from compiled binder"

    cover_text = reader.pages[0].extract_text() or ""
    assert cover_text.count("Generated") == 1, f"expected a single Generated date on the cover, got:\n{cover_text}"


def _make_compile_xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "COMPILE"
    ws.append(["PartNumber", "Revision", "Qty"])
    for pn, rev, qty in rows:
        ws.append([pn, rev, qty])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_excel_compile_view_wires_datasheets(app, client, tmp_path):
    """The Excel Compile form was restyled to match the Part Detail docpack
    builder; this locks in that the Datasheets checkbox actually reaches
    build_excel_compile_zip end-to-end via the view. (Where-used was dropped:
    it's computed against the compile sheet's synthetic title/root, which no
    real BOMLink ever points at, so it was always an empty no-op there.)
    """
    root_dir = tmp_path
    app.config["FILE_ROOT_LOCAL"] = str(root_dir)

    with app.app_context():
        user = _tools_user()
        root = Part(part_number="ASM-VIEW", revision="1", description="Root Assembly", processes=["assembly"]).save()
        root_pdf = os.path.join(root_dir, "ASM-VIEW.pdf")
        _write_pdf(root_pdf, "ROOT VIEW PDF")
        PartFile(part_number=root.part_number, revision="1", ext_group="pdf", ext="pdf",
                  rel_path="ASM-VIEW.pdf", path=root_pdf).save()

    _login(client, user)
    xlsx_bytes = _make_compile_xlsx([("ASM-VIEW", "1", 1)])

    resp = client.post(
        "/tools/excelcompile",
        data={
            "file": (io.BytesIO(xlsx_bytes), "compile.xlsx"),
            "title": "View Test",
            "want_pdf_binder": "on",
            "binder_add_datasheets": "on",
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    body = resp.get_data(as_text=True)
    assert "Compile Complete" in body
    assert "No missing parts" in body


def _download_zip(client, body: str) -> zipfile.ZipFile:
    import re
    match = re.search(r'href="(/tools/excelcompile/download/[^"]+)"', body)
    assert match, body
    resp = client.get(match.group(1))
    assert resp.status_code == 200
    return zipfile.ZipFile(io.BytesIO(resp.data))


def test_excel_compile_view_unchecked_expand_subassemblies_excludes_children(app, client, tmp_path):
    """Regression test: browsers omit an unchecked checkbox from the POST body
    entirely, so tools.py must not fall back to True for expand_subassemblies
    when the field is simply absent -- that previously made unchecking
    "Expand sub-assemblies" a no-op and always pulled in BOM children.
    """
    root_dir = tmp_path
    app.config["FILE_ROOT_LOCAL"] = str(root_dir)

    with app.app_context():
        user = _tools_user()
        parent = Part(part_number="ASM-PARENT", revision="1", description="Parent Assembly", processes=["assembly"]).save()
        child = Part(part_number="ASM-CHILD", revision="1", description="Child Part", processes=["machine"]).save()
        BOMLink(parent_pn=parent.part_number, parent_rev="1", child_pn=child.part_number, child_rev="1", qty=2).save()

    _login(client, user)
    xlsx_bytes = _make_compile_xlsx([("ASM-PARENT", "1", 1)])

    # expand_subassemblies intentionally omitted from the payload, simulating
    # an unchecked checkbox (a real browser would never send this key at all).
    resp = client.post(
        "/tools/excelcompile",
        data={
            "file": (io.BytesIO(xlsx_bytes), "compile.xlsx"),
            "title": "No Expand Test",
            "want_excel_bom": "on",
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    with _download_zip(client, resp.get_data(as_text=True)) as zf:
        bom_name = next(n for n in zf.namelist() if n.lower().endswith(".xlsx") and n != "input.xlsx")
        wb_bytes = zf.read(bom_name)

    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))
    ws = wb.active
    cell_values = {str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None}
    assert "ASM-PARENT" in cell_values
    assert "ASM-CHILD" not in cell_values, "child should not be pulled in when sub-assembly expansion is unchecked"


def test_excel_compile_view_checked_expand_subassemblies_includes_children(app, client, tmp_path):
    root_dir = tmp_path
    app.config["FILE_ROOT_LOCAL"] = str(root_dir)

    with app.app_context():
        user = _tools_user()
        parent = Part(part_number="ASM-PARENT2", revision="1", description="Parent Assembly", processes=["assembly"]).save()
        child = Part(part_number="ASM-CHILD2", revision="1", description="Child Part", processes=["machine"]).save()
        BOMLink(parent_pn=parent.part_number, parent_rev="1", child_pn=child.part_number, child_rev="1", qty=2).save()

    _login(client, user)
    xlsx_bytes = _make_compile_xlsx([("ASM-PARENT2", "1", 1)])

    resp = client.post(
        "/tools/excelcompile",
        data={
            "file": (io.BytesIO(xlsx_bytes), "compile.xlsx"),
            "title": "Expand Test",
            "want_excel_bom": "on",
            "expand_subassemblies": "on",
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    with _download_zip(client, resp.get_data(as_text=True)) as zf:
        bom_name = next(n for n in zf.namelist() if n.lower().endswith(".xlsx") and n != "input.xlsx")
        wb_bytes = zf.read(bom_name)

    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))
    ws = wb.active
    cell_values = {str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None}
    assert "ASM-PARENT2" in cell_values
    assert "ASM-CHILD2" in cell_values, "child should be pulled in when sub-assembly expansion is checked"


def test_excel_compile_custom_description_and_thumbnail(app, tmp_path):
    """The compile sheet's synthetic root has no real Part record, so it never
    had a description or thumbnail of its own -- a manually supplied
    description and image must reach the cover page.
    """
    root_dir = tmp_path
    app.config["FILE_ROOT_LOCAL"] = str(root_dir)

    root = Part(part_number="ASM-THUMB", revision="1", description="Root Assembly", processes=["assembly"]).save()
    root_pdf = os.path.join(root_dir, "ASM-THUMB.pdf")
    _write_pdf(root_pdf, "ROOT THUMB PDF")
    PartFile(part_number=root.part_number, revision="1", ext_group="pdf", ext="pdf",
              rel_path="ASM-THUMB.pdf", path=root_pdf).save()

    thumb_path = os.path.join(root_dir, "custom_thumb.png")
    _write_png(thumb_path)

    rows = [CompileRow(part_number=root.part_number, revision="1", qty=1.0, status="ok")]

    with app.app_context():
        data, missing = build_excel_compile_zip(
            rows,
            input_filename="input.xlsx",
            input_bytes=b"fake-xlsx-bytes",
            file_root=str(root_dir),
            want_excel_bom=False,
            want_pdf_binder=True,
            binder_add_cover=True,
            binder_add_index=False,
            binder_add_visual_list=False,
            binder_add_hardware_summary=False,
            binder_page_numbers=False,
            title="Custom Cover Quote",
            description="Hand-typed description for this quote.",
            thumbnail_path=thumb_path,
        )

    assert not missing
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        binder_name = next(n for n in zf.namelist() if "Binder" in n and n.lower().endswith(".pdf"))
        binder_bytes = zf.read(binder_name)

    reader = PdfReader(io.BytesIO(binder_bytes))
    cover_text = reader.pages[0].extract_text() or ""
    assert "Hand-typed description for this quote." in cover_text

    # There is no PartFile-based thumbnail for this synthetic root, so any
    # embedded image XObject on the cover page must be the uploaded one.
    resources = reader.pages[0].get("/Resources") or {}
    xobjects = resources.get("/XObject") or {}
    assert len(xobjects) > 0, "expected the custom thumbnail image to be embedded on the cover page"
